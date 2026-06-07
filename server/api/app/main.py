from __future__ import annotations
import json
import os
import shutil
import time
import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Annotated, Any

import requests
from fastapi import BackgroundTasks, Body, Depends, FastAPI, File, Form, Header, HTTPException, Request, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, EmailStr, Field
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .config import get_settings
from .database import Base, SessionLocal, engine, get_db
from .domain import (
    build_style_code,
    deserialize_traits,
    dumps_json,
    job_result_image_url,
    loads_json,
    serialize_traits,
)
from .models import (
    Booking,
    EventLog,
    Favorite,
    Merchant,
    NailStyleAsset,
    SessionToken,
    Store,
    StoreStyleListing,
    Style,
    StyleLifecycleRequest,
    StyleMetricsDaily,
    TryOnRecord,
    TrendPost,
    TrendRecommendation,
    TrendTopic,
    TryOnJob,
    User,
    HandImage,
)
from .logging_utils import log_event_json, setup_logging
from .security import create_access_token, hash_password, safe_decode_access_token, verify_password
from .services import (
    OpenClawCliAnalyzer,
    check_xhs_collection_status,
    collect_xiaohongshu_notes,
    generate_tryon_image,
    is_unusable_xhs_text,
    is_valid_nail_post,
    is_verified_xhs_post,
    sanitize_image_url,
)


ROLE_CUSTOMER = "customer"
ROLE_PLATFORM_ADMIN = "platform_admin"
ROLE_MERCHANT_ADMIN = "merchant_admin"
ROLE_MERCHANT_STAFF = "merchant_staff"
ADMIN_ROLES = {ROLE_PLATFORM_ADMIN, ROLE_MERCHANT_ADMIN, ROLE_MERCHANT_STAFF}
STYLE_ACTIVE_STATUSES = {"active"}
LEGACY_DEMO_STYLE_CODES = {"rose-mist", "tea-amber", "jade-ink"}
LEGACY_DEMO_STORE_CODES = {"s1", "s2", "s3"}
LEGACY_DEMO_TOPIC_KEYS = {"xh-rose-french", "xh-jade-modern", "xh-heavy-cat-eye"}

EVENT_TO_METRIC_FIELD = {
    "style_impression": "impressions",
    "style_click": "clicks",
    "style_favorite": "favorites",
    "tryon_start": "tryon_starts",
    "tryon_complete": "tryon_completes",
    "booking_create": "booking_creates",
    "booking_confirm": "booking_confirms",
}

ADMIN_WEB_DIR = Path(__file__).parent / "admin_web"
XHS_ACCOUNT_DIST_DIR = Path(
    os.getenv("XHS_ACCOUNT_DIST_DIR", str(Path(get_settings().data_dir) / "XHS_ALL_IN_ONE/frontend/dist"))
)
XHS_ACCOUNT_UPSTREAM = os.getenv("XHS_ACCOUNT_UPSTREAM", "http://172.17.0.1:8090")

settings = get_settings()
app_logger = setup_logging(settings.logs_dir)
app = FastAPI(title=settings.app_name)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"] if settings.allowed_origins_list == ["*"] else settings.allowed_origins_list,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
if ADMIN_WEB_DIR.exists():
    app.mount("/admin/static", StaticFiles(directory=ADMIN_WEB_DIR), name="admin-static")
app.mount("/files", StaticFiles(directory=get_settings().data_dir), name="data-files")


def _xhs_index_html() -> HTMLResponse:
    index_file = XHS_ACCOUNT_DIST_DIR / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=404, detail="XHS account matrix frontend is not deployed")
    html = index_file.read_text(encoding="utf-8")
    html = html.replace('src="/assets/', 'src="/xhs-account/assets/')
    html = html.replace('href="/assets/', 'href="/xhs-account/assets/')
    html = html.replace('href="/favicon.svg"', 'href="/xhs-account/favicon.svg"')
    return HTMLResponse(html)


def _proxy_xhs_response(path: str, request: Request, body: bytes) -> Response:
    upstream_url = f"{XHS_ACCOUNT_UPSTREAM.rstrip('/')}/api/{path}"
    excluded_headers = {"host", "content-length", "connection", "accept-encoding"}
    headers = {key: value for key, value in request.headers.items() if key.lower() not in excluded_headers}
    try:
        upstream_response = requests.request(
            request.method,
            upstream_url,
            params=list(request.query_params.multi_items()),
            data=body,
            headers=headers,
            timeout=60,
        )
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"XHS account matrix is unavailable: {exc}") from exc
    response_headers = {
        key: value
        for key, value in upstream_response.headers.items()
        if key.lower() not in {"content-length", "connection", "content-encoding", "transfer-encoding"}
    }
    return Response(
        content=upstream_response.content,
        status_code=upstream_response.status_code,
        headers=response_headers,
        media_type=upstream_response.headers.get("content-type"),
    )


@app.api_route("/xhs-account/api/{path:path}", methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"])
async def xhs_account_api_proxy(path: str, request: Request):
    return _proxy_xhs_response(path, request, await request.body())


@app.get("/xhs-account/assets/{path:path}")
def xhs_account_assets(path: str):
    asset_file = (XHS_ACCOUNT_DIST_DIR / "assets" / path).resolve()
    assets_dir = (XHS_ACCOUNT_DIST_DIR / "assets").resolve()
    if not str(asset_file).startswith(str(assets_dir)) or not asset_file.exists():
        raise HTTPException(status_code=404, detail="Asset not found")
    if asset_file.suffix == ".js":
        script = asset_file.read_text(encoding="utf-8")
        script = script.replace('baseURL:"/api"', 'baseURL:"/xhs-account/api"')
        script = script.replace('fetch("/api/', 'fetch("/xhs-account/api/')
        return Response(script, media_type="application/javascript")
    return FileResponse(asset_file)


@app.get("/xhs-account/favicon.svg")
def xhs_account_favicon():
    favicon = XHS_ACCOUNT_DIST_DIR / "favicon.svg"
    if not favicon.exists():
        raise HTTPException(status_code=404, detail="Favicon not found")
    return FileResponse(favicon)


@app.get("/xhs-account")
@app.get("/xhs-account/{path:path}")
def xhs_account_frontend(path: str = ""):
    return _xhs_index_html()


class RegisterRequest(BaseModel):
    name: str
    email: EmailStr
    password: str


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class BookingRequest(BaseModel):
    storeId: str
    styleId: str
    slot: str
    name: str
    phone: str
    note: str = ""


class CreateTryOnJobRequest(BaseModel):
    styleId: str
    sourceImageKey: str | None = None
    selectedLength: str = "natural_short"
    selectedShape: str = "squoval"


class RerenderTryOnJobRequest(BaseModel):
    selectedLength: str | None = None
    selectedShape: str | None = None


class WorkerProgressRequest(BaseModel):
    stage: str
    progress: int


class WorkerCompleteRequest(BaseModel):
    resultImageKey: str
    detectedTraits: dict[str, str] | None = None


class WorkerFailRequest(BaseModel):
    errorCode: str
    errorMessage: str


class SyncTryOnRequest(BaseModel):
    handId: str | None = None
    handImageId: int | None = None
    styleId: int
    selectedLength: str = "natural_short"
    selectedShape: str = "squoval"


class TrackEventRequest(BaseModel):
    eventName: str
    styleId: str | None = None
    storeId: str | None = None
    deviceId: str | None = None
    sourcePage: str | None = None
    sourceChannel: str | None = None
    sessionId: str | None = None
    payload: dict[str, Any] | None = None
    occurredAt: datetime | None = None


class AdminStyleCreateRequest(BaseModel):
    code: str | None = None
    name: str
    vibe: str
    price: str = ""
    nailType: str
    skinTone: str
    tags: list[str] = Field(default_factory=list)
    colors: list[str] = Field(default_factory=list)
    status: str = "draft"


class AdminStyleStatusRequest(BaseModel):
    status: str


class AdminStyleUpdateRequest(BaseModel):
    name: str | None = None
    vibe: str | None = None
    nailType: str | None = None
    skinTone: str | None = None
    tags: list[str] | None = None
    colors: list[str] | None = None
    status: str | None = None


class MerchantListingUpdateRequest(BaseModel):
    price: str | None = None
    inventoryCount: int | None = None
    status: str | None = None


class MerchantStoreUpdateRequest(BaseModel):
    slots: list[str] | None = None
    isAcceptingBookings: bool | None = None


class MerchantLifecycleRequestCreate(BaseModel):
    styleId: str
    requestedAction: str
    reason: str = ""
    storeId: str | None = None


class ReviewRequest(BaseModel):
    reviewNote: str = ""


class TrendImportRequest(BaseModel):
    title: str
    clusterLabel: str
    summary: str
    communityHeatScore: float
    targetStyleId: str | None = None
    recommendationType: str
    candidateName: str | None = None


class TrendCollectRequest(BaseModel):
    keywords: list[str]
    maxPostsPerKeyword: int = 6
    headless: bool = True


class TrendAnalyzeRequest(BaseModel):
    topicIds: list[int]


class XhsStorageStateRequest(BaseModel):
    storageState: dict[str, Any]


def utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def ensure_directories() -> None:
    Path(settings.data_dir).mkdir(parents=True, exist_ok=True)
    settings.uploads_dir.mkdir(parents=True, exist_ok=True)
    settings.results_dir.mkdir(parents=True, exist_ok=True)
    settings.static_dir.mkdir(parents=True, exist_ok=True)
    settings.static_styles_dir.mkdir(parents=True, exist_ok=True)
    settings.static_hands_dir.mkdir(parents=True, exist_ok=True)
    settings.logs_dir.mkdir(parents=True, exist_ok=True)


def public_file_url(local_path: str | None) -> str | None:
    if not local_path:
        return None
    path = Path(local_path)
    try:
        relative = path.relative_to(Path(settings.data_dir))
    except ValueError:
        return None
    return f"{settings.public_base_url.rstrip('/')}/files/{relative.as_posix()}"


def tryon_result_url(filename: str) -> str:
    return f"{settings.public_base_url.rstrip('/')}/files/results/{filename}"


@app.middleware("http")
async def request_logging_middleware(request: Request, call_next):
    request_id = uuid.uuid4().hex[:12]
    started_at = time.perf_counter()
    try:
        response = await call_next(request)
    except Exception:
        duration_ms = int((time.perf_counter() - started_at) * 1000)
        log_event_json(
            app_logger,
            "request_failed",
            requestId=request_id,
            method=request.method,
            path=request.url.path,
            query=request.url.query,
            durationMs=duration_ms,
            client=request.client.host if request.client else "",
        )
        raise

    duration_ms = int((time.perf_counter() - started_at) * 1000)
    response.headers["X-Request-Id"] = request_id
    log_event_json(
        app_logger,
        "request_completed",
        requestId=request_id,
        method=request.method,
        path=request.url.path,
        statusCode=response.status_code,
        durationMs=duration_ms,
        client=request.client.host if request.client else "",
    )
    return response


def get_style_asset_by_code(db: Session, style_code: str) -> NailStyleAsset | None:
    return db.scalar(select(NailStyleAsset).where(NailStyleAsset.style_code == style_code))


def get_style_asset_by_sequence(db: Session, style_id: int) -> NailStyleAsset | None:
    return db.scalar(select(NailStyleAsset).where(NailStyleAsset.id == style_id))


def next_numeric_code(prefix: str, latest_id: int | None) -> str:
    next_id = 1 if latest_id is None else latest_id + 1
    return f"{prefix}-{next_id:04d}"


def parse_prefixed_id(value: str, prefix: str) -> int:
    if not value.startswith(prefix):
        raise HTTPException(status_code=404, detail="resource not found")
    try:
        return int(value.removeprefix(prefix))
    except ValueError as exc:
        raise HTTPException(status_code=404, detail="resource not found") from exc


def user_response(user: User) -> dict[str, Any]:
    return {
        "name": user.name,
        "email": user.email,
        "preferences": [item for item in user.preferences.split(",") if item],
        "role": user.role,
        "merchantId": user.merchant.code if user.merchant else None,
        "managedStoreId": user.managed_store_code,
    }


def style_to_dict(style: Style, asset: NailStyleAsset | None = None) -> dict[str, Any]:
    image_url = public_file_url(asset.local_image_path) if asset and asset.local_image_path else (asset.enhanced_url if asset else None)
    return {
        "id": style.code,
        "name": style.name,
        "vibe": style.vibe,
        "price": style.price,
        "nailType": style.nail_type,
        "skinTone": style.skin_tone,
        "tags": loads_json(style.tags_json, []),
        "colors": loads_json(style.colors_json, []),
        "status": style.status,
        "imageUrl": image_url,
        "tryOnStyleId": asset.id if asset else None,
    }


def style_to_dict_with_db(db: Session, style: Style) -> dict[str, Any]:
    return style_to_dict(style, get_style_asset_by_code(db, style.code))


def has_real_style_image(db: Session, style: Style) -> bool:
    asset = get_style_asset_by_code(db, style.code)
    return bool(asset and (asset.local_image_path or asset.enhanced_url))


def is_real_workspace_style(db: Session, style: Style | None) -> bool:
    return bool(style and style.code not in LEGACY_DEMO_STYLE_CODES and has_real_style_image(db, style))


def prioritize_home_styles(db: Session, styles: list[Style]) -> list[Style]:
    def sort_key(style: Style) -> tuple[int, int, str]:
        asset = get_style_asset_by_code(db, style.code)
        has_real_image = bool(asset and (asset.local_image_path or asset.enhanced_url))
        sequence = asset.sequence_no if asset else 9999
        return (0 if has_real_image else 1, sequence, style.code)

    return sorted(styles, key=sort_key)


def store_to_dict(store: Store) -> dict[str, Any]:
    return {
        "id": store.code,
        "name": store.name,
        "distance": store.distance,
        "priceBand": store.price_band,
        "score": store.score,
        "slots": loads_json(store.slots_json, []),
        "openHours": store.open_hours,
        "artists": store.artists,
        "works": store.works,
        "isAcceptingBookings": store.is_accepting_bookings,
        "merchantId": store.merchant.code if store.merchant else None,
    }


def listing_to_dict(listing: StoreStyleListing) -> dict[str, Any]:
    return {
        "id": listing.id,
        "storeId": listing.store_code,
        "styleId": listing.style_code,
        "price": listing.price,
        "inventoryCount": listing.inventory_count,
        "status": listing.status,
        "publishedAt": listing.published_at.isoformat() if listing.published_at else None,
    }


def listing_to_dict_with_style(db: Session, listing: StoreStyleListing) -> dict[str, Any]:
    payload = listing_to_dict(listing)
    style = find_style(db, listing.style_code)
    payload["style"] = style_to_dict_with_db(db, style) if style else None
    return payload


def recommendation_to_dict(rec: TrendRecommendation) -> dict[str, Any]:
    candidate_payload = loads_json(rec.candidate_payload_json, None)
    return {
        "id": rec.recommendation_code,
        "type": rec.recommendation_type,
        "targetStyleId": rec.target_style_code,
        "targetStoreId": rec.target_store_code,
        "candidateName": rec.candidate_name,
        "triggerReason": rec.trigger_reason,
        "communityEvidence": rec.community_evidence,
        "inAppEvidence": rec.in_app_evidence,
        "confidenceScore": rec.confidence_score,
        "actionText": rec.action_text,
        "prerequisites": rec.prerequisites,
        "candidatePayload": candidate_payload,
        "imageUrl": candidate_payload.get("imageUrl") if isinstance(candidate_payload, dict) else None,
        "status": rec.status,
        "createdAt": rec.created_at.isoformat(),
        "reviewedAt": rec.reviewed_at.isoformat() if rec.reviewed_at else None,
    }


def lifecycle_request_to_dict(req: StyleLifecycleRequest) -> dict[str, Any]:
    style_name = None
    store_name = None
    if req.style_code:
        style_name = req.style_code
    if req.store_code:
        store_name = req.store_code
    return {
        "id": req.request_code,
        "requestedByUserId": req.requested_by_user_id,
        "merchantId": req.merchant_id,
        "storeId": req.store_code,
        "storeName": store_name,
        "styleId": req.style_code,
        "styleName": style_name,
        "requestedAction": req.requested_action,
        "reason": req.reason,
        "status": req.status,
        "reviewNote": req.review_note,
        "reviewedByUserId": req.reviewed_by_user_id,
        "reviewedAt": req.reviewed_at.isoformat() if req.reviewed_at else None,
        "createdAt": req.created_at.isoformat(),
    }


def booking_response(booking: Booking) -> dict[str, Any]:
    return {
        "id": f"booking-{booking.id:03d}",
        "status": booking.status,
        "storeId": booking.store_id,
        "storeName": booking.store_name,
        "styleId": booking.style_id,
        "styleName": booking.style_name,
        "slot": booking.slot,
        "price": booking.price,
        "name": booking.name,
        "phone": booking.phone,
        "note": booking.note,
        "createdAt": booking.created_at.isoformat(),
        "confirmedAt": booking.confirmed_at.isoformat() if booking.confirmed_at else None,
    }


def try_on_job_response(job: TryOnJob) -> dict[str, Any]:
    return {
        "id": job.job_code,
        "userId": f"user-{job.user_id:03d}",
        "styleId": job.style_id,
        "styleName": job.style_name,
        "sourceImageKey": job.source_image_key,
        "status": job.status,
        "stage": job.stage,
        "progress": job.progress,
        "selectedLength": job.selected_length,
        "selectedShape": job.selected_shape,
        "resultImageKey": job.result_image_key,
        "detectedTraits": deserialize_traits(job.detected_traits),
        "errorCode": job.error_code,
        "errorMessage": job.error_message,
        "createdAt": job.created_at.isoformat(),
        "updatedAt": job.updated_at.isoformat(),
        "completedAt": job.completed_at.isoformat() if job.completed_at else None,
    }


def extract_job_code_from_result_url(result_url: str) -> str | None:
    marker = "/api/try-on/jobs/"
    if marker not in result_url:
        return None
    try:
        tail = result_url.split(marker, 1)[1]
        return tail.split("/", 1)[0]
    except Exception:
        return None


def queue_try_on_job(background_tasks: BackgroundTasks, job_code: str) -> None:
    background_tasks.add_task(process_tryon_job, job_code)


def process_tryon_job(job_code: str) -> None:
    started_at = time.perf_counter()
    with SessionLocal() as db:
        job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code))
        if not job:
            return
        try:
            job.status = "processing"
            job.stage = "preparing"
            job.progress = 15
            job.claimed_by_worker = True
            job.error_code = None
            job.error_message = None
            job.updated_at = utcnow()
            db.add(job)
            db.commit()

            asset = get_style_asset_by_code(db, job.style_id)
            if not asset:
                raise RuntimeError("style asset not found")

            source_path = settings.uploads_dir / job.source_image_key
            if not source_path.exists():
                raise RuntimeError("source image not found")

            job.stage = "loading_image"
            job.progress = 35
            job.updated_at = utcnow()
            db.add(job)
            db.commit()

            style_path = ensure_style_image_local(db, asset)
            db.commit()

            job.stage = "rendering"
            job.progress = 70
            job.updated_at = utcnow()
            db.add(job)
            db.commit()

            result_filename = f"{job.job_code}.png"
            result_path = settings.results_dir / result_filename
            success, message = generate_tryon_image(str(source_path), str(style_path), str(result_path))
            if not success:
                raise RuntimeError(message)

            duration_ms = int((time.perf_counter() - started_at) * 1000)
            job.result_image_key = result_filename
            job.status = "completed"
            job.stage = "completed"
            job.progress = 100
            job.completed_at = utcnow()
            job.updated_at = utcnow()
            db.add(job)
            db.add(
                TryOnRecord(
                    user_id=job.user_id,
                    hand_image_id=None,
                    nail_style_asset_id=asset.id,
                    result_url=tryon_result_url(result_filename),
                    source="async-job",
                    duration_ms=duration_ms,
                    selected_length=job.selected_length,
                    selected_shape=job.selected_shape,
                )
            )
            log_event(db, "tryon_complete", user_id=job.user_id, style_id=job.style_id, source_page="tryon_async")
            db.commit()
            log_event_json(
                app_logger,
                "tryon_async_completed",
                jobId=job.job_code,
                userId=job.user_id,
                styleId=job.style_id,
                durationMs=duration_ms,
                resultImageKey=result_filename,
            )
        except Exception as exc:
            db.rollback()
            job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code))
            if not job:
                return
            job.status = "failed"
            job.stage = "failed"
            job.progress = 100
            job.error_code = "async_tryon_failed"
            job.error_message = str(exc)
            job.claimed_by_worker = False
            job.updated_at = utcnow()
            db.add(job)
            db.commit()
            log_event_json(
                app_logger,
                "tryon_async_failed",
                jobId=job_code,
                reason=str(exc),
            )


def try_on_record_response(db: Session, record: TryOnRecord) -> dict[str, Any]:
    asset = db.get(NailStyleAsset, record.nail_style_asset_id)
    return {
        "id": f"tryon-record-{record.id:04d}",
        "jobId": extract_job_code_from_result_url(record.result_url),
        "resultUrl": record.result_url,
        "durationMs": record.duration_ms,
        "styleName": asset.display_name if asset else "",
        "styleId": asset.style_code if asset else "",
        "source": record.source,
        "selectedLength": record.selected_length,
        "selectedShape": record.selected_shape,
        "createdAt": record.created_at.isoformat(),
    }


def fetch_binary_source(path_or_url: str) -> bytes:
    if path_or_url.startswith("http://") or path_or_url.startswith("https://"):
        response = requests.get(path_or_url, timeout=30)
        response.raise_for_status()
        return response.content
    return Path(path_or_url).read_bytes()


def image_suffix_from_source(path_or_url: str | None, default: str = ".png") -> str:
    if not path_or_url:
        return default
    suffix = Path(path_or_url.split("?")[0]).suffix.lower()
    return suffix if suffix in {".png", ".jpg", ".jpeg", ".webp"} else default


def ensure_local_binary(source: str, target: Path) -> Path:
    if target.exists() and target.stat().st_size > 1000:
        return target
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(fetch_binary_source(source))
    return target


def ensure_style_image_local(db: Session, asset: NailStyleAsset) -> Path:
    if asset.local_image_path:
        existing = Path(asset.local_image_path)
        if existing.exists() and existing.stat().st_size > 1000:
            return existing
    source = asset.enhanced_url or asset.original_url
    if not source:
        raise HTTPException(status_code=400, detail="style image source missing")
    target = settings.static_styles_dir / f"style_{asset.sequence_no:02d}{image_suffix_from_source(source)}"
    local_path = ensure_local_binary(source, target)
    asset.local_image_path = str(local_path)
    asset.updated_at = utcnow()
    db.add(asset)
    return local_path


def next_style_asset_sequence(db: Session) -> int:
    last = db.scalar(select(func.max(NailStyleAsset.sequence_no)))
    return int(last or 0) + 1


def ensure_style_asset_record(db: Session, style: Style) -> NailStyleAsset:
    asset = get_style_asset_by_code(db, style.code)
    if asset:
        return asset
    tags = loads_json(style.tags_json, [])
    asset = NailStyleAsset(
        style_code=style.code,
        display_name=style.name,
        sequence_no=next_style_asset_sequence(db),
        original_url=None,
        enhanced_url=None,
        local_image_path=None,
        category=tags[0] if tags else style.nail_type,
        color_tone="#F8C7D6",
        tags_json=style.tags_json,
        price=style.price,
        popularity=0,
        updated_at=utcnow(),
    )
    db.add(asset)
    db.flush()
    return asset


def ensure_hand_image_local(db: Session, hand: HandImage) -> Path:
    if hand.local_path:
        existing = Path(hand.local_path)
        if existing.exists() and existing.stat().st_size > 1000:
            return existing
    source = hand.image_url
    if not source:
        raise HTTPException(status_code=400, detail="hand image source missing")
    suffix = image_suffix_from_source(source, default=".jpg")
    target = settings.static_hands_dir / f"{hand.hand_code}{suffix}"
    local_path = ensure_local_binary(source, target)
    hand.local_path = str(local_path)
    db.add(hand)
    return local_path


def create_session_for_user(db: Session, user: User) -> str:
    token = create_access_token(user.id, user.email)
    db.add(SessionToken(user_id=user.id, token=token))
    db.commit()
    return token


def get_current_user(
    authorization: Annotated[str | None, Header()] = None,
    db: Session = Depends(get_db),
) -> User:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="missing bearer token")
    token = authorization.removeprefix("Bearer ").strip()
    payload = safe_decode_access_token(token)
    if not payload:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="invalid or expired token")
    session_row = db.scalar(select(SessionToken).where(SessionToken.token == token))
    if not session_row:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="invalid or expired token")
    user = db.get(User, int(payload["sub"]))
    if not user or not user.is_active:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="user not found")
    return user


def get_admin_user(user: User = Depends(get_current_user)) -> User:
    if user.role not in ADMIN_ROLES:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="admin access required")
    return user


def require_platform_admin(user: User = Depends(get_admin_user)) -> User:
    if user.role != ROLE_PLATFORM_ADMIN:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="platform admin access required")
    return user


def require_merchant_user(user: User = Depends(get_admin_user)) -> User:
    if user.role not in {ROLE_MERCHANT_ADMIN, ROLE_MERCHANT_STAFF}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="merchant access required")
    return user


def require_worker(
    x_worker_token: Annotated[str | None, Header()] = None,
    authorization: Annotated[str | None, Header()] = None,
) -> None:
    token = x_worker_token or ""
    if not token and authorization and authorization.startswith("Bearer "):
        token = authorization.removeprefix("Bearer ").strip()
    if token != settings.worker_token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="invalid worker token")


def find_style(db: Session, style_code: str, include_inactive: bool = True) -> Style | None:
    stmt = select(Style).where(Style.code == style_code)
    if not include_inactive:
        stmt = stmt.where(Style.status.in_(tuple(STYLE_ACTIVE_STATUSES)))
    return db.scalar(stmt)


def find_store(db: Session, store_code: str) -> Store | None:
    return db.scalar(select(Store).where(Store.code == store_code))


def match_style(style: Style, query: str) -> bool:
    query = query.strip().lower()
    if not query:
        return True
    fields = [style.name, style.vibe, style.nail_type, style.skin_tone, *loads_json(style.tags_json, [])]
    return any(query in field.lower() for field in fields)


def next_job_code(db: Session) -> str:
    last = db.scalar(select(TryOnJob).order_by(TryOnJob.id.desc()))
    return next_numeric_code("tryon", last.id if last else None)


def next_request_code(db: Session) -> str:
    last = db.scalar(select(StyleLifecycleRequest).order_by(StyleLifecycleRequest.id.desc()))
    return next_numeric_code("req", last.id if last else None)


def next_recommendation_code(db: Session) -> str:
    last = db.scalar(select(TrendRecommendation).order_by(TrendRecommendation.id.desc()))
    return next_numeric_code("rec", last.id if last else None)


def build_topic_key(keyword: str) -> str:
    return build_style_code(f"xhs-{keyword}")[:40]


def load_trend_post_meta(post: TrendPost) -> dict[str, Any]:
    meta = loads_json(post.extracted_tags_json, {})
    return meta if isinstance(meta, dict) else {"tags": meta if isinstance(meta, list) else []}


def is_verified_trend_post_meta(meta: dict[str, Any]) -> bool:
    return meta.get("verified") is True


def trend_post_score(post: TrendPost) -> int:
    return int(post.like_count or 0) + int(post.collect_count or 0) * 2 + int(post.comment_count or 0) * 3


def trend_post_to_dict(post: TrendPost) -> dict[str, Any]:
    meta = load_trend_post_meta(post)
    return {
        "postId": post.post_id,
        "topicId": post.topic_id,
        "url": post.url,
        "title": post.title,
        "author": post.author,
        "imageUrl": meta.get("imageUrl"),
        "tags": meta.get("tags", []),
        "keyword": meta.get("keyword"),
        "verified": is_verified_trend_post_meta(meta),
        "likeCount": post.like_count,
        "collectCount": post.collect_count,
        "commentCount": post.comment_count,
        "engagementScore": trend_post_score(post),
        "publishedAt": post.published_at.isoformat() if post.published_at else None,
    }


def is_usable_trend_post(post: TrendPost) -> bool:
    payload = trend_post_to_dict(post)
    payload["imageUrl"] = sanitize_image_url(payload.get("imageUrl"))
    return payload.get("verified") is True and is_verified_xhs_post(payload, payload.get("keyword"))


def build_community_trend_dashboard(db: Session) -> dict[str, Any]:
    topics = db.scalars(
        select(TrendTopic).order_by(TrendTopic.community_heat_score.desc(), TrendTopic.last_seen_at.desc()).limit(18)
    ).all()
    if not topics:
        return {"topics": [], "clusters": [], "hotPosts": [], "stats": {"topics": 0, "clusters": 0, "posts": 0}}

    topic_ids = [item.id for item in topics]
    posts = [item for item in db.scalars(select(TrendPost).where(TrendPost.topic_id.in_(topic_ids))).all() if is_usable_trend_post(item)]
    posts_by_topic: dict[int, list[TrendPost]] = defaultdict(list)
    for post in posts:
        posts_by_topic[post.topic_id].append(post)

    topic_cards = []
    cluster_map: dict[str, dict[str, Any]] = {}
    for topic in topics:
        topic_posts = sorted(posts_by_topic.get(topic.id, []), key=trend_post_score, reverse=True)
        if not topic_posts and is_unusable_xhs_text(topic.title):
            continue
        lead_post = topic_posts[0] if topic_posts else None
        topic_cards.append(
            {
                "topicKey": topic.topic_key,
                "title": topic.title,
                "clusterLabel": topic.cluster_label,
                "summary": topic.summary,
                "heatScore": round(topic.community_heat_score, 1),
                "evidenceCount": topic.evidence_count,
                "lastSeenAt": topic.last_seen_at.isoformat() if topic.last_seen_at else None,
                "leadPost": trend_post_to_dict(lead_post) if lead_post else None,
            }
        )
        cluster_key = (topic.cluster_label or topic.title or topic.topic_key).strip()
        bucket = cluster_map.setdefault(
            cluster_key,
            {
                "label": cluster_key,
                "heatScore": 0.0,
                "topicCount": 0,
                "postCount": 0,
                "leadPost": None,
                "topicTitles": [],
            },
        )
        bucket["heatScore"] += float(topic.community_heat_score or 0.0)
        bucket["topicCount"] += 1
        bucket["postCount"] += len(topic_posts)
        bucket["topicTitles"].append(topic.title)
        if lead_post and (bucket["leadPost"] is None or trend_post_score(lead_post) > bucket["leadPost"]["engagementScore"]):
            bucket["leadPost"] = trend_post_to_dict(lead_post)

    clusters = sorted(
        (
            {
                **item,
                "heatScore": round(item["heatScore"], 1),
                "topicTitles": item["topicTitles"][:4],
            }
            for item in cluster_map.values()
        ),
        key=lambda item: (item["heatScore"], item["postCount"]),
        reverse=True,
    )[:8]

    hot_posts = sorted(posts, key=trend_post_score, reverse=True)
    return {
        "topics": topic_cards[:8],
        "clusters": clusters,
        "hotPosts": [trend_post_to_dict(item) for item in hot_posts[:8]],
        "stats": {
            "topics": len(topic_cards),
            "clusters": len(clusters),
            "posts": len(posts),
        },
    }


def build_product_signal_label(summary: dict[str, Any]) -> str:
    funnel = summary["funnel"]
    scores = summary["scores"]
    if funnel["bookingCreates"] >= 2 and funnel["favorites"] >= 3:
        return "收藏和预约同步上升"
    if funnel["tryonStarts"] >= 3 and funnel["clicks"] >= 4:
        return "试戴兴趣显著"
    if scores["healthLabel"] == "optimize_candidate":
        return "有兴趣但还没转化"
    if scores["healthLabel"] == "delist_candidate":
        return "曝光有了，点击偏弱"
    return "站内表现稳定"


def build_product_trend_dashboard(db: Session) -> dict[str, Any]:
    style_rows = [style for style in db.scalars(select(Style).order_by(Style.code)).all() if is_real_workspace_style(db, style)]
    items = []
    for style in style_rows:
        style_payload = style_to_dict_with_db(db, style)
        analytics = build_style_trend_summary(db, style.code)
        funnel = analytics["funnel"]
        scores = analytics["scores"]
        items.append(
            {
                "style": style_payload,
                "styleId": style.code,
                "styleName": style.name,
                "imageUrl": style_payload.get("imageUrl"),
                "impressions": funnel["impressions"],
                "clicks": funnel["clicks"],
                "favorites": funnel["favorites"],
                "tryonStarts": funnel["tryonStarts"],
                "tryonCompletes": funnel["tryonCompletes"],
                "bookingCreates": funnel["bookingCreates"],
                "bookingConfirms": funnel["bookingConfirms"],
                "clickThroughRate": funnel["clickThroughRate"],
                "favoriteRate": funnel["favoriteRate"],
                "bookingConversionRate": funnel["bookingConversionRate"],
                "communityHeatScore": scores["communityHeatScore"],
                "inAppInterestScore": scores["inAppInterestScore"],
                "compositeRecommendationScore": scores["compositeRecommendationScore"],
                "healthLabel": scores["healthLabel"],
                "signalLabel": build_product_signal_label(analytics),
            }
        )

    hot_styles = sorted(
        items,
        key=lambda item: (
            item["compositeRecommendationScore"],
            item["favorites"],
            item["tryonStarts"],
            item["bookingCreates"],
        ),
        reverse=True,
    )[:8]
    conversion_leaders = sorted(
        [item for item in items if item["clicks"] > 0],
        key=lambda item: (item["bookingConversionRate"], item["favoriteRate"], item["clicks"]),
        reverse=True,
    )[:6]
    funnel_counts = aggregate_event_counts(db, days=7)
    return {
        "hotStyles": hot_styles,
        "conversionLeaders": conversion_leaders,
        "stats": {
            "styles": len(items),
            "impressions": funnel_counts["style_impression"],
            "favorites": funnel_counts["style_favorite"],
            "tryonStarts": funnel_counts["tryon_start"],
            "bookingCreates": funnel_counts["booking_create"],
        },
    }


def build_ai_curated_community_posts(recommendations: list[TrendRecommendation]) -> list[dict[str, Any]]:
    curated: list[dict[str, Any]] = []
    seen: set[str] = set()
    for rec in recommendations:
        candidate_payload = loads_json(rec.candidate_payload_json, {}) if rec.candidate_payload_json else {}
        if not isinstance(candidate_payload, dict):
            continue
        for post in candidate_payload.get("sourcePosts", []):
            if not isinstance(post, dict) or post.get("verified") is not True or not is_verified_xhs_post(post, None):
                continue
            like_count = int(post.get("likeCount", 0) or 0)
            collect_count = int(post.get("collectCount", 0) or 0)
            comment_count = int(post.get("commentCount", 0) or 0)
            engagement_score = like_count + collect_count * 2 + comment_count * 3
            if engagement_score <= 0 or engagement_score > 500000:
                continue
            image_url = sanitize_image_url(post.get("imageUrl"))
            if not image_url:
                continue
            key = post.get("url") or post.get("postId") or post.get("title")
            if not key or key in seen:
                continue
            seen.add(key)
            curated.append(
                {
                    "postId": post.get("postId"),
                    "title": post.get("title"),
                    "author": post.get("author"),
                    "imageUrl": image_url,
                    "tags": post.get("tags", []),
                    "engagementLabel": "OpenClaw 已筛选",
                    "recommendedBy": rec.candidate_name or rec.recommendation_code,
                }
            )
            if len(curated) >= 8:
                return curated
    return curated


def save_xiaohongshu_collection(db: Session, groups: list[dict[str, Any]]) -> list[TrendTopic]:
    topics: list[TrendTopic] = []
    for group in groups:
        keyword = group["keyword"]
        topic_key = build_topic_key(keyword)
        topic = db.scalar(select(TrendTopic).where(TrendTopic.topic_key == topic_key))
        if not topic:
            topic = TrendTopic(
                topic_key=topic_key,
                platform="xiaohongshu",
                title=group["topicTitle"],
                cluster_label=group["clusterLabel"],
                summary=group["summary"],
                community_heat_score=group["communityHeatScore"],
                evidence_count=len(group["posts"]),
                last_seen_at=utcnow(),
            )
            db.add(topic)
            db.flush()
        else:
            topic.title = group["topicTitle"]
            topic.cluster_label = group["clusterLabel"]
            topic.summary = group["summary"]
            topic.community_heat_score = group["communityHeatScore"]
            topic.evidence_count = len(group["posts"])
            topic.last_seen_at = utcnow()
            db.add(topic)

        for post_payload in group["posts"]:
            post = db.scalar(select(TrendPost).where(TrendPost.post_id == post_payload["postId"]))
            post_meta = {
                "tags": post_payload.get("tags", []),
                "imageUrl": post_payload.get("imageUrl"),
                "keyword": keyword,
                "sourceNoteId": post_payload["postId"],
                "verified": bool(post_payload.get("verified")),
            }
            if not post:
                post = TrendPost(
                    topic_id=topic.id,
                    post_id=post_payload["postId"],
                    url=post_payload["url"],
                    title=post_payload["title"],
                    author=post_payload.get("author") or "",
                    like_count=int(post_payload.get("likeCount", 0)),
                    collect_count=int(post_payload.get("collectCount", 0)),
                    comment_count=int(post_payload.get("commentCount", 0)),
                    published_at=utcnow(),
                    extracted_tags_json=dumps_json(post_meta),
                )
            else:
                post.topic_id = topic.id
                post.url = post_payload["url"]
                post.title = post_payload["title"]
                post.author = post_payload.get("author") or ""
                post.like_count = int(post_payload.get("likeCount", 0))
                post.collect_count = int(post_payload.get("collectCount", 0))
                post.comment_count = int(post_payload.get("commentCount", 0))
                post.extracted_tags_json = dumps_json(post_meta)
            db.add(post)
        topics.append(topic)
    db.commit()
    return topics


def build_recommendation_posts(db: Session, topic: TrendTopic) -> list[dict[str, Any]]:
    posts = [
        item
        for item in db.scalars(select(TrendPost).where(TrendPost.topic_id == topic.id).order_by(TrendPost.like_count.desc())).all()
        if is_usable_trend_post(item)
    ]
    payload = []
    for post in posts[:5]:
        meta = load_trend_post_meta(post)
        payload.append(
            {
                "postId": post.post_id,
                "url": post.url,
                "title": post.title,
                "author": post.author,
                "imageUrl": meta.get("imageUrl"),
                "tags": meta.get("tags", []),
                "likeCount": post.like_count,
                "collectCount": post.collect_count,
                "commentCount": post.comment_count,
            }
        )
    return payload


def build_trend_recommendations(db: Session, topics: list[TrendTopic]) -> int:
    analyzer = OpenClawCliAnalyzer()
    created = 0
    last_rec = db.scalar(select(TrendRecommendation).order_by(TrendRecommendation.id.desc()))
    next_rec_number = (last_rec.id if last_rec else 0) + 1
    for topic in topics:
        existing = db.scalar(select(TrendRecommendation).where(TrendRecommendation.trigger_reason == topic.summary))
        if existing:
            continue
        posts = build_recommendation_posts(db, topic)
        try:
            summary = analyzer.summarize(
                title=topic.title,
                cluster_label=topic.cluster_label,
                summary=topic.summary,
                heat=topic.community_heat_score,
                evidence_count=topic.evidence_count,
                posts=posts,
            )
        except Exception as exc:
            log_event_json(app_logger, "trend_collect_analysis_failed", topicKey=topic.topic_key, error=str(exc))
            continue
        db.add(
            TrendRecommendation(
                recommendation_code=f"rec-{next_rec_number:04d}",
                recommendation_type=summary["recommendation_type"],
                target_style_code=None,
                target_store_code=None,
                candidate_name=summary["candidate_name"],
                trigger_reason=summary["trigger_reason"],
                community_evidence=summary["community_evidence"],
                in_app_evidence=summary["in_app_evidence"],
                confidence_score=summary["confidence_score"],
                action_text=summary["action_text"],
                prerequisites=summary["prerequisites"],
                candidate_payload_json=dumps_json(summary["candidate_payload"]),
                status="pending",
            )
        )
        db.flush()
        next_rec_number += 1
        created += 1
    db.commit()
    return created


def compute_health_label(impressions: int, clicks: int, bookings: int) -> str:
    if impressions < 20 and clicks < 3:
        return "insufficient_distribution"
    ctr = clicks / impressions if impressions else 0
    if impressions >= 20 and ctr < 0.08:
        return "delist_candidate"
    if clicks >= 8 and bookings == 0:
        return "optimize_candidate"
    return "healthy"


def refresh_metric_scores(db: Session, row: StyleMetricsDaily) -> None:
    ctr = row.clicks / row.impressions if row.impressions else 0.0
    favorite_rate = row.favorites / row.clicks if row.clicks else 0.0
    tryon_rate = row.tryon_starts / row.clicks if row.clicks else 0.0
    booking_rate = row.booking_creates / row.clicks if row.clicks else 0.0
    community_heat_confidence = db.scalar(
        select(func.max(TrendRecommendation.confidence_score)).where(TrendRecommendation.target_style_code == row.style_id)
    )
    row.community_heat_score = round(float((community_heat_confidence or 0.0) * 100), 2)
    row.in_app_interest_score = round((ctr * 40 + favorite_rate * 20 + tryon_rate * 20 + booking_rate * 20) * 100, 2)
    row.composite_recommendation_score = round((row.community_heat_score * 0.55) + (row.in_app_interest_score * 0.45), 2)
    row.health_label = compute_health_label(row.impressions, row.clicks, row.booking_creates)
    row.updated_at = utcnow()


def upsert_style_metric(db: Session, style_code: str | None, event_name: str, occurred_at: datetime) -> None:
    if not style_code or event_name not in EVENT_TO_METRIC_FIELD:
        return
    metric_day = occurred_at.date()
    row = db.scalar(select(StyleMetricsDaily).where(StyleMetricsDaily.style_id == style_code, StyleMetricsDaily.metric_date == metric_day))
    if not row:
        row = StyleMetricsDaily(style_id=style_code, metric_date=metric_day)
        db.add(row)
        db.flush()
    field_name = EVENT_TO_METRIC_FIELD[event_name]
    setattr(row, field_name, getattr(row, field_name) + 1)
    refresh_metric_scores(db, row)


def log_event(
    db: Session,
    event_name: str,
    *,
    user_id: int | None = None,
    device_id: str | None = None,
    style_id: str | None = None,
    store_id: str | None = None,
    source_page: str | None = None,
    source_channel: str | None = None,
    session_id: str | None = None,
    payload: dict[str, Any] | None = None,
    occurred_at: datetime | None = None,
) -> str:
    event_time = occurred_at or utcnow()
    event_id = f"evt-{uuid.uuid4().hex[:20]}"
    db.add(
        EventLog(
            event_id=event_id,
            event_name=event_name,
            user_id=user_id,
            device_id=device_id,
            style_id=style_id,
            store_id=store_id,
            source_page=source_page,
            source_channel=source_channel,
            session_id=session_id,
            payload_json=dumps_json(payload) if payload else None,
            occurred_at=event_time,
        )
    )
    upsert_style_metric(db, style_id, event_name, event_time)
    return event_id


def log_impressions(db: Session, user: User | None, styles: list[Style], source_page: str) -> None:
    for style in styles:
        log_event(db, "style_impression", user_id=user.id if user else None, style_id=style.code, source_page=source_page)
    if styles:
        db.commit()


def latest_style_metrics(db: Session, style_code: str) -> StyleMetricsDaily | None:
    return db.scalar(
        select(StyleMetricsDaily)
        .where(StyleMetricsDaily.style_id == style_code)
        .order_by(StyleMetricsDaily.metric_date.desc())
    )


def aggregate_event_counts(db: Session, *, style_id: str | None = None, days: int = 7) -> dict[str, int]:
    since = utcnow() - timedelta(days=days)
    stmt = select(EventLog.event_name, func.count(EventLog.id)).where(EventLog.occurred_at >= since).group_by(EventLog.event_name)
    if style_id:
        stmt = stmt.where(EventLog.style_id == style_id)
    rows = db.execute(stmt).all()
    counts = {name: count for name, count in rows}
    for event_name in EVENT_TO_METRIC_FIELD:
        counts.setdefault(event_name, 0)
    return counts


def build_style_trend_summary(db: Session, style_code: str) -> dict[str, Any]:
    daily_rows = db.scalars(
        select(StyleMetricsDaily)
        .where(StyleMetricsDaily.style_id == style_code, StyleMetricsDaily.metric_date >= date.today() - timedelta(days=6))
        .order_by(StyleMetricsDaily.metric_date.asc())
    ).all()
    counts = aggregate_event_counts(db, style_id=style_code, days=7)
    impressions = counts["style_impression"]
    clicks = counts["style_click"]
    favorites = counts["style_favorite"]
    tryon_starts = counts["tryon_start"]
    tryon_completes = counts["tryon_complete"]
    booking_creates = counts["booking_create"]
    booking_confirms = counts["booking_confirm"]
    latest = latest_style_metrics(db, style_code)
    total_heat = latest.community_heat_score if latest else 0.0
    health_label = latest.health_label if latest else compute_health_label(impressions, clicks, booking_creates)
    return {
        "styleId": style_code,
        "funnel": {
            "impressions": impressions,
            "clicks": clicks,
            "favorites": favorites,
            "tryonStarts": tryon_starts,
            "tryonCompletes": tryon_completes,
            "bookingCreates": booking_creates,
            "bookingConfirms": booking_confirms,
            "clickThroughRate": round(clicks / impressions, 4) if impressions else 0.0,
            "favoriteRate": round(favorites / clicks, 4) if clicks else 0.0,
            "tryonStartRate": round(tryon_starts / clicks, 4) if clicks else 0.0,
            "tryonCompleteRate": round(tryon_completes / tryon_starts, 4) if tryon_starts else 0.0,
            "bookingConversionRate": round(booking_creates / clicks, 4) if clicks else 0.0,
            "dealConversionRate": round(booking_confirms / clicks, 4) if clicks else 0.0,
        },
        "scores": {
            "communityHeatScore": round(total_heat, 2),
            "inAppInterestScore": round(latest.in_app_interest_score if latest else 0.0, 2),
            "compositeRecommendationScore": round(latest.composite_recommendation_score if latest else 0.0, 2),
            "healthLabel": health_label,
        },
        "daily": [
            {
                "date": row.metric_date.isoformat(),
                "impressions": row.impressions,
                "clicks": row.clicks,
                "favorites": row.favorites,
                "tryonStarts": row.tryon_starts,
                "tryonCompletes": row.tryon_completes,
                "bookingCreates": row.booking_creates,
                "bookingConfirms": row.booking_confirms,
                "communityHeatScore": row.community_heat_score,
                "compositeRecommendationScore": row.composite_recommendation_score,
                "healthLabel": row.health_label,
            }
            for row in daily_rows
        ],
    }


def merchant_store_scope(user: User) -> list[str]:
    if user.role == ROLE_PLATFORM_ADMIN:
        return []
    if user.managed_store_code:
        return [user.managed_store_code]
    return []


def require_store_access(user: User, store_code: str) -> None:
    if user.role == ROLE_PLATFORM_ADMIN:
        return
    if user.managed_store_code != store_code:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="store access denied")


def get_or_create_listing(db: Session, store_code: str, style_code: str, price: str) -> StoreStyleListing:
    listing = db.scalar(select(StoreStyleListing).where(StoreStyleListing.store_code == store_code, StoreStyleListing.style_code == style_code))
    if listing:
        return listing
    listing = StoreStyleListing(
        store_code=store_code,
        style_code=style_code,
        price=price,
        inventory_count=6,
        status="draft",
        updated_at=utcnow(),
    )
    db.add(listing)
    db.flush()
    return listing


def derive_hot_keywords(styles: list[Style]) -> list[str]:
    seen: list[str] = []
    for style in styles:
        for tag in loads_json(style.tags_json, []):
            normalized = str(tag).strip()
            if normalized and normalized not in seen:
                seen.append(normalized)
        for token in str(style.vibe or "").replace("，", ",").split(","):
            normalized = token.strip()
            if normalized and normalized not in seen:
                seen.append(normalized)
        if len(seen) >= 8:
            break
    return seen[:8]


def cleanup_legacy_demo_data(db: Session) -> None:
    db.query(TrendRecommendation).filter(
        (TrendRecommendation.target_style_code.in_(tuple(LEGACY_DEMO_STYLE_CODES))) |
        (TrendRecommendation.trigger_reason.like("%小红书近期通勤显白法式笔记互动显著上升%")) |
        (TrendRecommendation.trigger_reason.like("%节气和婚礼场景带动新中式玉石纹样热度上升%")) |
        (TrendRecommendation.trigger_reason.like("%重闪高饱和猫眼内容仍有曝光%")) |
        (TrendRecommendation.candidate_name.like("%当前笔记暂时无法浏览%")) |
        (TrendRecommendation.trigger_reason.like("%当前笔记暂时无法浏览%")) |
        (TrendRecommendation.community_evidence.like("%当前笔记暂时无法浏览%"))
    ).delete(synchronize_session=False)
    demo_topic_ids = [item.id for item in db.scalars(select(TrendTopic).where(TrendTopic.topic_key.in_(tuple(LEGACY_DEMO_TOPIC_KEYS)))).all()]
    if demo_topic_ids:
        db.query(TrendPost).filter(TrendPost.topic_id.in_(demo_topic_ids)).delete(synchronize_session=False)
        db.query(TrendTopic).filter(TrendTopic.id.in_(demo_topic_ids)).delete(synchronize_session=False)
    db.query(StoreStyleListing).filter(
        (StoreStyleListing.store_code.in_(tuple(LEGACY_DEMO_STORE_CODES))) |
        (StoreStyleListing.style_code.in_(tuple(LEGACY_DEMO_STYLE_CODES)))
    ).delete(synchronize_session=False)
    db.query(StyleLifecycleRequest).filter(
        (StyleLifecycleRequest.store_code.in_(tuple(LEGACY_DEMO_STORE_CODES))) |
        (StyleLifecycleRequest.style_code.in_(tuple(LEGACY_DEMO_STYLE_CODES)))
    ).delete(synchronize_session=False)
    db.query(StyleMetricsDaily).filter(StyleMetricsDaily.style_id.in_(tuple(LEGACY_DEMO_STYLE_CODES))).delete(synchronize_session=False)
    db.query(Favorite).filter(Favorite.style_id.in_(tuple(LEGACY_DEMO_STYLE_CODES))).delete(synchronize_session=False)
    db.query(Booking).filter(
        (Booking.style_id.in_(tuple(LEGACY_DEMO_STYLE_CODES))) |
        (Booking.store_id.in_(tuple(LEGACY_DEMO_STORE_CODES)))
    ).delete(synchronize_session=False)
    db.query(TryOnJob).filter(TryOnJob.style_id.in_(tuple(LEGACY_DEMO_STYLE_CODES))).delete(synchronize_session=False)
    db.query(Store).filter(Store.code.in_(tuple(LEGACY_DEMO_STORE_CODES))).delete(synchronize_session=False)
    db.query(Style).filter(Style.code.in_(tuple(LEGACY_DEMO_STYLE_CODES))).delete(synchronize_session=False)


IMPORT_STYLE_NAMES = [
    "法式简约", "星空渐变", "樱花粉", "经典红", "裸色优雅",
    "闪钻奢华", "莫兰迪绿", "雾霾蓝", "焦糖棕", "玫瑰金",
    "大理石纹理", "奶油白", "薰衣草紫", "蜜桃橘", "薄荷绿",
    "深海蓝", "酒红丝绒", "豆沙粉", "香槟金", "银河流星",
    "暗黑系", "马卡龙", "蜜糖裸", "彩虹渐变", "珍珠白",
]
IMPORT_STYLE_CATEGORIES = ["纯色", "渐变", "法式", "闪粉", "手绘", "猫眼", "晕染", "大理石"]
IMPORT_STYLE_COLORS = [
    "#ffe4e1", "#6a5acd", "#ffb7c5", "#dc143c", "#deb887",
    "#ffd700", "#98fb98", "#87ceeb", "#d2691e", "#e8b4b8",
    "#f5deb3", "#b0c4de", "#dda0dd", "#f0e68c", "#20b2aa",
]
IMPORT_SKIN_TONES = ["白皙", "自然", "小麦", "白皙", "自然"]
IMPORT_HAND_TYPES = ["纤细", "标准", "圆润", "标准", "纤细"]


def workbook_path() -> Path:
    return Path.cwd() / "命题三美甲评测数据（对外版）.xlsx"


def seed_tryon_assets(db: Session) -> None:
    if db.scalar(select(func.count(NailStyleAsset.id))) and db.scalar(select(func.count(HandImage.id))):
        return

    rows_styles: list[tuple[int, str | None, str | None]] = []
    rows_hands: list[str] = []
    path = workbook_path()
    if path.exists():
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=True, data_only=True)
            style_sheet = wb["款式图"]
            hand_sheet = wb["手图"]
            for idx, row in enumerate(style_sheet.iter_rows(values_only=True), start=1):
                if idx == 1:
                    continue
                rows_styles.append((int(row[0]), row[1], row[2]))
            seen_hand_urls: set[str] = set()
            for idx, row in enumerate(hand_sheet.iter_rows(values_only=True), start=1):
                if idx == 1 or not row or not row[0]:
                    continue
                hand_url = str(row[0])
                if hand_url in seen_hand_urls:
                    continue
                seen_hand_urls.add(hand_url)
                rows_hands.append(hand_url)
        except Exception:
            rows_styles = []
            rows_hands = []

    if not rows_styles or not rows_hands:
        return

    for idx, original_url, enhanced_url in rows_styles:
        style_code = f"style-{idx:02d}"
        asset = db.scalar(select(NailStyleAsset).where(NailStyleAsset.style_code == style_code))
        if not asset:
            asset = NailStyleAsset(
                style_code=style_code,
                display_name=f"款式 {idx:02d}",
                sequence_no=idx,
                original_url=original_url,
                enhanced_url=enhanced_url,
                local_image_path=None,
                category="",
                color_tone="",
                tags_json=dumps_json([]),
                price="",
                popularity=0,
                updated_at=utcnow(),
            )
            db.add(asset)

        style = db.scalar(select(Style).where(Style.code == style_code))
        if not style:
            db.add(
                Style(
                    code=style_code,
                    name=asset.display_name if asset else f"款式 {idx:02d}",
                    vibe="",
                    price="",
                    nail_type="",
                    skin_tone="",
                    tags_json=dumps_json([]),
                    colors_json=dumps_json([]),
                    status="active",
                    updated_at=utcnow(),
                )
            )

    for idx, hand_url in enumerate(rows_hands, start=1):
        hand_code = f"hand_{idx:02d}"
        hand = db.scalar(select(HandImage).where(HandImage.hand_code == hand_code))
        if not hand:
            db.add(
                HandImage(
                    hand_code=hand_code,
                    image_url=hand_url,
                    local_path=None,
                    source_type="preset",
                    skin_tone="",
                    hand_type="",
                )
            )

    db.commit()


@app.on_event("startup")
def on_startup() -> None:
    ensure_directories()
    Base.metadata.create_all(bind=engine)
    with SessionLocal() as db:
        cleanup_legacy_demo_data(db)
        db.commit()
        seed_tryon_assets(db)
    log_event_json(
        app_logger,
        "api_started",
        appName=settings.app_name,
        dataDir=settings.data_dir,
        publicBaseUrl=settings.public_base_url,
    )


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/admin", response_class=HTMLResponse)
def admin_index() -> HTMLResponse:
    index_file = ADMIN_WEB_DIR / "index.html"
    if index_file.exists():
        return HTMLResponse(index_file.read_text(encoding="utf-8"))
    return HTMLResponse("<html><body><h1>Nail Mind Admin</h1><p>admin web not found</p></body></html>")


@app.post("/api/auth/register")
def register(payload: RegisterRequest, db: Session = Depends(get_db)) -> dict[str, Any]:
    if len(payload.password) < 6 or not payload.name.strip():
        raise HTTPException(status_code=400, detail="name, email and password(>=6) are required")
    if db.scalar(select(User).where(User.email == payload.email.lower())):
        raise HTTPException(status_code=409, detail="email already registered")
    user = User(
        email=payload.email.lower(),
        password_hash=hash_password(payload.password),
        name=payload.name.strip(),
        role=ROLE_CUSTOMER,
        preferences="显白,法式",
        style_preferences="显白、法式",
        notifications="试戴完成、预约提醒、活动通知",
        privacy="手部照片仅用于试戴与订单关联",
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    token = create_session_for_user(db, user)
    return {"token": token, "user": user_response(user)}


@app.post("/api/auth/login")
def login(payload: LoginRequest, db: Session = Depends(get_db)) -> dict[str, Any]:
    user = db.scalar(select(User).where(User.email == payload.email.lower()))
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="invalid email or password")
    token = create_session_for_user(db, user)
    return {"token": token, "user": user_response(user)}


@app.get("/api/auth/me")
def auth_me(user: User = Depends(get_current_user)) -> dict[str, Any]:
    return {"user": user_response(user)}


@app.post("/api/auth/logout")
def logout(
    authorization: Annotated[str | None, Header()] = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    token = authorization.removeprefix("Bearer ").strip()
    session_row = db.scalar(select(SessionToken).where(SessionToken.token == token, SessionToken.user_id == user.id))
    if session_row:
        db.delete(session_row)
        db.commit()
    return {"status": "logged_out"}


@app.get("/api/home")
def home(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    items = [
        style for style in db.scalars(select(Style).where(Style.status.in_(tuple(STYLE_ACTIVE_STATUSES))).order_by(Style.code)).all()
        if has_real_style_image(db, style)
    ]
    ranked_items = prioritize_home_styles(db, items)
    recommended_items = ranked_items[:6]
    hot_items = ranked_items[6:15] or ranked_items[:9]
    log_impressions(db, user, recommended_items + hot_items, "home")
    return {
        "hotKeywords": derive_hot_keywords(ranked_items),
        "recommended": [style_to_dict_with_db(db, style) for style in recommended_items],
        "hot": [style_to_dict_with_db(db, style) for style in hot_items],
    }


@app.get("/api/styles")
def styles(tag: str | None = None, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    items = [
        style for style in db.scalars(select(Style).where(Style.status.in_(tuple(STYLE_ACTIVE_STATUSES))).order_by(Style.code)).all()
        if has_real_style_image(db, style)
    ]
    if tag:
        items = [style for style in items if match_style(style, tag)]
    log_impressions(db, user, items, "styles")
    return {"items": [style_to_dict_with_db(db, style) for style in items]}


@app.get("/api/styles/search")
def style_search(q: str = "", user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    query = q.strip()
    items = [
        style for style in db.scalars(select(Style).where(Style.status.in_(tuple(STYLE_ACTIVE_STATUSES))).order_by(Style.code)).all()
        if has_real_style_image(db, style)
    ]
    result = [style for style in items if match_style(style, query)] if query else items
    log_event(db, "search_submit", user_id=user.id, source_page="styles_search", payload={"query": query, "resultCount": len(result)})
    db.commit()
    log_impressions(db, user, result, "styles_search")
    return {"query": query, "items": [style_to_dict_with_db(db, style) for style in result]}


@app.get("/api/styles/{style_id}")
def style_detail(style_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    style = find_style(db, style_id, include_inactive=False)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    favorited = db.scalar(select(Favorite).where(Favorite.user_id == user.id, Favorite.style_id == style_id)) is not None
    log_event(db, "style_click", user_id=user.id, style_id=style_id, source_page="style_detail")
    db.commit()
    return {"style": style_to_dict_with_db(db, style), "userCases": "", "canFavorite": True, "canTryOn": True, "favorited": favorited}


@app.get("/api/favorites")
def favorites(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    favorite_ids = {row.style_id for row in db.scalars(select(Favorite).where(Favorite.user_id == user.id)).all()}
    items = [
        style for style in db.scalars(select(Style).where(Style.code.in_(favorite_ids)).order_by(Style.code)).all()
        if has_real_style_image(db, style)
    ] if favorite_ids else []
    return {"items": [style_to_dict_with_db(db, style) for style in items]}


@app.post("/api/favorites/{style_id}")
def add_favorite(style_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    if not find_style(db, style_id, include_inactive=False):
        raise HTTPException(status_code=404, detail="style not found")
    existing = db.scalar(select(Favorite).where(Favorite.user_id == user.id, Favorite.style_id == style_id))
    if not existing:
        db.add(Favorite(user_id=user.id, style_id=style_id))
    log_event(db, "style_favorite", user_id=user.id, style_id=style_id, source_page="favorites")
    db.commit()
    return {"styleId": style_id, "favorited": True}


@app.delete("/api/favorites/{style_id}")
def remove_favorite(style_id: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    existing = db.scalar(select(Favorite).where(Favorite.user_id == user.id, Favorite.style_id == style_id))
    if existing:
        db.delete(existing)
        db.commit()
    return {"styleId": style_id, "favorited": False}


@app.post("/api/events")
def track_event(payload: TrackEventRequest, user: User | None = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, str]:
    event_id = log_event(
        db,
        payload.eventName,
        user_id=user.id if user else None,
        device_id=payload.deviceId,
        style_id=payload.styleId,
        store_id=payload.storeId,
        source_page=payload.sourcePage,
        source_channel=payload.sourceChannel,
        session_id=payload.sessionId,
        payload=payload.payload,
        occurred_at=payload.occurredAt,
    )
    db.commit()
    return {"eventId": event_id}


@app.get("/api/tryon/hand-images")
def list_hand_images(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    items = db.scalars(select(HandImage).order_by(HandImage.created_at.desc())).all()
    return {
        "hands": [
            {
                "id": item.hand_code,
                "dbId": item.id,
                "imageUrl": public_file_url(item.local_path) if item.local_path else item.image_url,
                "sourceType": item.source_type,
                "skinTone": item.skin_tone,
                "handType": item.hand_type,
            }
            for item in items
        ],
        "total": len(items),
    }


@app.post("/api/tryon/upload-hand")
def upload_hand_image(file: UploadFile = File(...), user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    suffix = Path(file.filename or "upload.png").suffix.lower() or ".png"
    if suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
        raise HTTPException(status_code=400, detail="unsupported image format")
    next_id = (db.scalar(select(func.max(HandImage.id))) or 0) + 1
    hand_code = f"user_{next_id:04d}"
    target_path = settings.uploads_dir / f"{hand_code}{suffix}"
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with target_path.open("wb") as output:
        shutil.copyfileobj(file.file, output)
    hand = HandImage(
        hand_code=hand_code,
        image_url=public_file_url(str(target_path)) or f"/files/uploads/{target_path.name}",
        local_path=str(target_path),
        source_type="user",
        skin_tone="auto",
        hand_type="auto",
    )
    db.add(hand)
    db.commit()
    db.refresh(hand)
    user.last_upload_key = hand.hand_code
    db.add(user)
    db.commit()
    return {"hand_id": hand.hand_code, "image_url": public_file_url(hand.local_path), "db_id": hand.id, "message": "上传成功"}


@app.post("/api/tryon/try-on")
def sync_try_on(payload: SyncTryOnRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    asset = db.get(NailStyleAsset, payload.styleId)
    if not asset:
        raise HTTPException(status_code=404, detail="style asset not found")

    hand: HandImage | None = None
    if payload.handImageId is not None:
        hand = db.get(HandImage, payload.handImageId)
    elif payload.handId:
        hand = db.scalar(select(HandImage).where(HandImage.hand_code == payload.handId))
    if not hand:
        raise HTTPException(status_code=404, detail="hand image not found")

    hand_name = hand.hand_code
    style_name = f"style_{asset.sequence_no:02d}"
    result_filename = f"{hand_name}+{style_name}+{payload.selectedLength}+{payload.selectedShape}.png"
    result_path = settings.results_dir / result_filename
    source = "bailian-cached" if result_path.exists() and result_path.stat().st_size > 1000 else ""
    started_at = datetime.now()

    if not source:
        style_path = ensure_style_image_local(db, asset)
        hand_path = ensure_hand_image_local(db, hand)
        db.commit()
        success, message = generate_tryon_image(str(hand_path), str(style_path), str(result_path))
        if not success:
            log_event_json(
                app_logger,
                "tryon_failed",
                userId=user.id,
                handId=hand.hand_code,
                styleId=asset.style_code,
                provider="dashscope",
                reason=message,
            )
            raise HTTPException(status_code=502, detail=f"large-model try-on failed: {message}")
        source = "bailian-live"

    duration_ms = int((datetime.now() - started_at).total_seconds() * 1000)
    result_url = tryon_result_url(result_filename)
    record = TryOnRecord(
        user_id=user.id,
        hand_image_id=hand.id,
        nail_style_asset_id=asset.id,
        result_url=result_url,
        source=source,
        duration_ms=duration_ms,
        selected_length=payload.selectedLength,
        selected_shape=payload.selectedShape,
    )
    db.add(record)
    log_event(
        db,
        "tryon_sync",
        user_id=user.id,
        style_id=asset.style_code,
        source_page="tryon",
        payload={"source": source, "durationMs": duration_ms},
    )
    db.commit()
    log_event_json(
        app_logger,
        "tryon_completed",
        userId=user.id,
        handId=hand.hand_code,
        styleId=asset.style_code,
        source=source,
        durationMs=duration_ms,
        selectedLength=payload.selectedLength,
        selectedShape=payload.selectedShape,
        resultFile=result_filename,
    )
    return {"result_url": result_url, "duration_ms": duration_ms, "style_name": asset.display_name, "source": source}


@app.get("/api/tryon/history")
def try_on_history(limit: int = 20, offset: int = 0, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    stmt = (
        select(TryOnRecord)
        .where(TryOnRecord.user_id == user.id)
        .order_by(TryOnRecord.created_at.desc())
        .offset(offset)
        .limit(limit)
    )
    items = db.scalars(stmt).all()
    total = db.scalar(select(func.count(TryOnRecord.id)).where(TryOnRecord.user_id == user.id)) or 0
    return {"items": [try_on_record_response(db, item) for item in items], "total": total}


@app.post("/api/try-on/uploads")
def tryon_upload(file: UploadFile = File(...), user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    suffix = Path(file.filename or "upload.jpg").suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".png"}:
        raise HTTPException(status_code=400, detail="only .jpg, .jpeg and .png files are supported")
    object_key = f"user-{user.id:03d}/{int(utcnow().timestamp() * 1_000_000)}{suffix}"
    target = settings.uploads_dir / object_key
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("wb") as output:
        shutil.copyfileobj(file.file, output)
    user.last_upload_key = object_key
    db.add(user)
    db.commit()
    return {"objectKey": object_key, "fileName": file.filename}


@app.get("/api/try-on/jobs")
def tryon_jobs(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    jobs = db.scalars(select(TryOnJob).where(TryOnJob.user_id == user.id).order_by(TryOnJob.created_at.desc())).all()
    return {"items": [try_on_job_response(job) for job in jobs]}


@app.post("/api/try-on/jobs")
def create_tryon_job(
    payload: CreateTryOnJobRequest,
    background_tasks: BackgroundTasks,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    style = find_style(db, payload.styleId, include_inactive=False)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    source_key = payload.sourceImageKey or user.last_upload_key
    if not source_key:
        raise HTTPException(status_code=400, detail="sourceImageKey is required")
    if not (settings.uploads_dir / source_key).exists():
        raise HTTPException(status_code=400, detail="source image not found")
    job = TryOnJob(
        job_code=next_job_code(db),
        user_id=user.id,
        style_id=style.code,
        style_name=style.name,
        source_image_key=source_key,
        status="queued",
        stage="queued",
        progress=5,
        selected_length=payload.selectedLength,
        selected_shape=payload.selectedShape,
        updated_at=utcnow(),
    )
    db.add(job)
    log_event(db, "tryon_start", user_id=user.id, style_id=style.code, source_page="tryon")
    db.commit()
    db.refresh(job)
    queue_try_on_job(background_tasks, job.job_code)
    return try_on_job_response(job)


@app.get("/api/try-on/jobs/{job_code}")
def get_tryon_job(job_code: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code, TryOnJob.user_id == user.id))
    if not job:
        raise HTTPException(status_code=404, detail="try-on job not found")
    return try_on_job_response(job)


@app.get("/api/try-on/jobs/{job_code}/result")
def get_tryon_result(job_code: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code, TryOnJob.user_id == user.id))
    if not job:
        raise HTTPException(status_code=404, detail="try-on job not found")
    payload = try_on_job_response(job)
    if job.status == "completed":
        payload["resultImageUrl"] = job_result_image_url(job.job_code)
    return payload


@app.get("/api/try-on/jobs/{job_code}/result-image")
def get_tryon_result_image(job_code: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> FileResponse:
    job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code, TryOnJob.user_id == user.id))
    if not job or not job.result_image_key:
        raise HTTPException(status_code=404, detail="result image not found")
    path = settings.results_dir / job.result_image_key
    if not path.exists():
        raise HTTPException(status_code=404, detail="result image not found")
    return FileResponse(path, media_type="image/png")


@app.post("/api/try-on/jobs/{job_code}/rerender")
def rerender_tryon_job(
    job_code: str,
    payload: RerenderTryOnJobRequest,
    background_tasks: BackgroundTasks,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code, TryOnJob.user_id == user.id))
    if not job:
        raise HTTPException(status_code=404, detail="try-on job not found")
    if payload.selectedLength:
        job.selected_length = payload.selectedLength
    if payload.selectedShape:
        job.selected_shape = payload.selectedShape
    job.status = "queued"
    job.stage = "queued"
    job.progress = 5
    job.error_code = None
    job.error_message = None
    job.completed_at = None
    job.updated_at = utcnow()
    db.add(job)
    db.commit()
    db.refresh(job)
    queue_try_on_job(background_tasks, job.job_code)
    return try_on_job_response(job)


@app.get("/api/stores")
def stores(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    active_styles = {style.code for style in db.scalars(select(Style).where(Style.status.in_(tuple(STYLE_ACTIVE_STATUSES)))).all()}
    store_rows = db.scalars(select(Store).order_by(Store.code)).all()
    visible = []
    for store in store_rows:
        listing_count = db.scalar(
            select(func.count(StoreStyleListing.id)).where(
                StoreStyleListing.store_code == store.code,
                StoreStyleListing.style_code.in_(active_styles),
                StoreStyleListing.status == "active",
            )
        )
        if listing_count:
            visible.append(store)
    return {"items": [store_to_dict(store) for store in visible]}


@app.get("/api/stores/{store_id}")
def store_detail(store_id: str, _: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    store = find_store(db, store_id)
    if not store:
        raise HTTPException(status_code=404, detail="store not found")
    return store_to_dict(store)


@app.get("/api/stores/{store_id}/slots")
def store_slots(store_id: str, _: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    store = find_store(db, store_id)
    if not store:
        raise HTTPException(status_code=404, detail="store not found")
    return {"storeId": store.code, "slots": loads_json(store.slots_json, [])}


@app.get("/api/bookings")
def bookings(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    items = db.scalars(select(Booking).where(Booking.user_id == user.id).order_by(Booking.created_at.desc())).all()
    return {"items": [booking_response(booking) for booking in items]}


@app.post("/api/bookings")
def create_booking(payload: BookingRequest, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    style = find_style(db, payload.styleId, include_inactive=False)
    store = find_store(db, payload.storeId)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    if not store:
        raise HTTPException(status_code=404, detail="store not found")
    slots = loads_json(store.slots_json, [])
    if payload.slot not in slots:
        raise HTTPException(status_code=400, detail="slot is not available for the selected store")
    listing = db.scalar(
        select(StoreStyleListing).where(
            StoreStyleListing.store_code == store.code,
            StoreStyleListing.style_code == style.code,
            StoreStyleListing.status == "active",
        )
    )
    if not listing or listing.inventory_count <= 0:
        raise HTTPException(status_code=400, detail="selected style is not available at this store")
    booking = Booking(
        user_id=user.id,
        status="pending_confirmation",
        store_id=store.code,
        store_name=store.name,
        style_id=style.code,
        style_name=style.name,
        slot=payload.slot,
        price=listing.price,
        name=payload.name,
        phone=payload.phone,
        note=payload.note,
    )
    db.add(booking)
    listing.inventory_count = max(listing.inventory_count - 1, 0)
    listing.updated_at = utcnow()
    log_event(db, "booking_create", user_id=user.id, style_id=style.code, store_id=store.code, source_page="booking")
    db.commit()
    db.refresh(booking)
    return booking_response(booking)


@app.get("/api/bookings/{booking_code}")
def get_booking(booking_code: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    booking_id = parse_prefixed_id(booking_code, "booking-")
    booking = db.scalar(select(Booking).where(Booking.id == booking_id, Booking.user_id == user.id))
    if not booking:
        raise HTTPException(status_code=404, detail="booking not found")
    return booking_response(booking)


@app.post("/api/bookings/{booking_code}/confirm")
def confirm_booking(booking_code: str, user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    booking_id = parse_prefixed_id(booking_code, "booking-")
    booking = db.scalar(select(Booking).where(Booking.id == booking_id, Booking.user_id == user.id))
    if not booking:
        raise HTTPException(status_code=404, detail="booking not found")
    booking.status = "confirmed"
    booking.confirmed_at = utcnow()
    db.add(booking)
    log_event(db, "booking_confirm", user_id=user.id, style_id=booking.style_id, store_id=booking.store_id, source_page="booking_detail")
    db.commit()
    db.refresh(booking)
    return booking_response(booking)


@app.get("/api/profile")
def profile(user: User = Depends(get_current_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    favorites_count = db.query(Favorite).filter(Favorite.user_id == user.id).count()
    booking_count = db.query(Booking).filter(Booking.user_id == user.id).count()
    tryon_count = db.query(TryOnJob).filter(TryOnJob.user_id == user.id).count()
    return {"profile": user_response(user), "favoritesCount": favorites_count, "bookingCount": booking_count, "tryOnCount": tryon_count}


@app.get("/api/settings")
def user_settings(user: User = Depends(get_current_user)) -> dict[str, Any]:
    return {
        "stylePreferences": user.style_preferences,
        "notifications": user.notifications,
        "privacy": user.privacy,
    }


@app.post("/admin/auth/login")
def admin_login(payload: LoginRequest, db: Session = Depends(get_db)) -> dict[str, Any]:
    user = db.scalar(select(User).where(User.email == payload.email.lower()))
    if not user or not verify_password(payload.password, user.password_hash) or user.role not in ADMIN_ROLES:
        raise HTTPException(status_code=401, detail="invalid admin credentials")
    token = create_session_for_user(db, user)
    return {"token": token, "user": user_response(user)}


@app.get("/admin/auth/me")
def admin_me(user: User = Depends(get_admin_user)) -> dict[str, Any]:
    return {"user": user_response(user)}


@app.post("/admin/auth/logout")
def admin_logout(
    authorization: Annotated[str | None, Header()] = None,
    user: User = Depends(get_admin_user),
    db: Session = Depends(get_db),
) -> dict[str, str]:
    token = authorization.removeprefix("Bearer ").strip()
    session_row = db.scalar(select(SessionToken).where(SessionToken.token == token, SessionToken.user_id == user.id))
    if session_row:
        db.delete(session_row)
        db.commit()
    return {"status": "logged_out"}


@app.get("/admin/styles")
def admin_styles(user: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    items = [style for style in db.scalars(select(Style).order_by(Style.code)).all() if style.code not in LEGACY_DEMO_STYLE_CODES]
    return {"items": [style_to_dict_with_db(db, style) for style in items], "role": user.role}


@app.post("/admin/styles")
def admin_create_style(payload: AdminStyleCreateRequest, _: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    code = payload.code or build_style_code(payload.name)
    if find_style(db, code):
        raise HTTPException(status_code=409, detail="style code already exists")
    style = Style(
        code=code,
        name=payload.name,
        vibe=payload.vibe,
        price=payload.price,
        nail_type=payload.nailType,
        skin_tone=payload.skinTone,
        tags_json=dumps_json(payload.tags),
        colors_json=dumps_json(payload.colors),
        status=payload.status,
        updated_at=utcnow(),
    )
    db.add(style)
    db.commit()
    db.refresh(style)
    return {"style": style_to_dict_with_db(db, style)}


@app.patch("/admin/styles/{style_id}")
def admin_update_style(
    style_id: str,
    payload: AdminStyleUpdateRequest,
    _: User = Depends(require_platform_admin),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    style = find_style(db, style_id)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    if payload.name is not None:
        style.name = payload.name
    if payload.vibe is not None:
        style.vibe = payload.vibe
    if payload.nailType is not None:
        style.nail_type = payload.nailType
    if payload.skinTone is not None:
        style.skin_tone = payload.skinTone
    if payload.tags is not None:
        style.tags_json = dumps_json(payload.tags)
    if payload.colors is not None:
        style.colors_json = dumps_json(payload.colors)
    if payload.status is not None:
        style.status = payload.status
    style.updated_at = utcnow()
    asset = ensure_style_asset_record(db, style)
    asset.display_name = style.name
    asset.tags_json = style.tags_json
    asset.category = (payload.tags or loads_json(style.tags_json, []))[0] if loads_json(style.tags_json, []) else style.nail_type
    asset.updated_at = utcnow()
    db.add(style)
    db.add(asset)
    db.commit()
    return {"style": style_to_dict_with_db(db, style)}


@app.get("/admin/styles/{style_id}")
def admin_style_detail(style_id: str, _: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    style = find_style(db, style_id)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    listings = db.scalars(select(StoreStyleListing).where(StoreStyleListing.style_code == style_id).order_by(StoreStyleListing.store_code)).all()
    return {
        "style": style_to_dict_with_db(db, style),
        "listings": [listing_to_dict(item) for item in listings],
        "analytics": build_style_trend_summary(db, style_id),
    }


@app.post("/admin/styles/{style_id}/status")
def admin_update_style_status(style_id: str, payload: AdminStyleStatusRequest, _: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    style = find_style(db, style_id)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    style.status = payload.status
    style.updated_at = utcnow()
    db.add(style)
    db.commit()
    return {"style": style_to_dict_with_db(db, style)}


@app.post("/admin/styles/{style_id}/image")
def admin_update_style_image(
    style_id: str,
    image: UploadFile = File(...),
    _: User = Depends(require_platform_admin),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    style = find_style(db, style_id)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    asset = ensure_style_asset_record(db, style)
    suffix = image_suffix_from_source(image.filename or "", default=".png")
    target = settings.static_styles_dir / f"style_{asset.sequence_no:02d}{suffix}"
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("wb") as file_obj:
        shutil.copyfileobj(image.file, file_obj)
    asset.local_image_path = str(target)
    asset.display_name = style.name
    asset.updated_at = utcnow()
    db.add(asset)
    db.commit()
    return {"style": style_to_dict_with_db(db, style)}


@app.get("/admin/stores")
def admin_stores(_: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    items = db.scalars(select(Store).order_by(Store.code)).all()
    payload = []
    for store in items:
        listings = db.query(StoreStyleListing).filter(StoreStyleListing.store_code == store.code).count()
        pending = db.query(StyleLifecycleRequest).filter(StyleLifecycleRequest.store_code == store.code, StyleLifecycleRequest.status == "pending").count()
        store_payload = store_to_dict(store)
        store_payload["listingCount"] = listings
        store_payload["pendingRequestCount"] = pending
        payload.append(store_payload)
    return {"items": payload}


@app.get("/admin/stores/{store_id}")
def admin_store_detail(store_id: str, user: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    require_store_access(user, store_id)
    store = find_store(db, store_id)
    if not store:
        raise HTTPException(status_code=404, detail="store not found")
    listings = db.scalars(select(StoreStyleListing).where(StoreStyleListing.store_code == store_id).order_by(StoreStyleListing.style_code)).all()
    bookings = db.scalars(select(Booking).where(Booking.store_id == store_id).order_by(Booking.created_at.desc()).limit(10)).all()
    return {
        "store": store_to_dict(store),
        "listings": [listing_to_dict(item) for item in listings],
        "recentBookings": [booking_response(item) for item in bookings],
    }


@app.get("/admin/requests")
def admin_requests(user: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    stmt = select(StyleLifecycleRequest).order_by(StyleLifecycleRequest.created_at.desc())
    if user.role in {ROLE_MERCHANT_ADMIN, ROLE_MERCHANT_STAFF} and user.managed_store_code:
        stmt = stmt.where(StyleLifecycleRequest.store_code == user.managed_store_code)
    items = db.scalars(stmt).all()
    payload = []
    for item in items:
        request_payload = lifecycle_request_to_dict(item)
        style = find_style(db, item.style_code) if item.style_code else None
        store = find_store(db, item.store_code) if item.store_code else None
        request_payload["styleName"] = style.name if style else item.style_code
        request_payload["storeName"] = store.name if store else item.store_code
        payload.append(request_payload)
    return {"items": payload}


def apply_lifecycle_request(db: Session, req: StyleLifecycleRequest) -> None:
    style = find_style(db, req.style_code)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    listing = None
    if req.store_code:
        store = find_store(db, req.store_code)
        if not store:
            raise HTTPException(status_code=404, detail="store not found")
        listing = get_or_create_listing(db, req.store_code, req.style_code, style.price)
    if req.requested_action == "launch":
        style.status = "active"
        if listing:
            listing.status = "active"
            listing.published_at = listing.published_at or utcnow()
    elif req.requested_action == "delist":
        if listing:
            listing.status = "inactive"
        else:
            style.status = "inactive"
    else:
        raise HTTPException(status_code=400, detail="unsupported request action")
    style.updated_at = utcnow()


@app.post("/admin/requests/{request_id}/approve")
def approve_request(request_id: str, payload: ReviewRequest, user: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    req = db.scalar(select(StyleLifecycleRequest).where(StyleLifecycleRequest.request_code == request_id))
    if not req:
        raise HTTPException(status_code=404, detail="request not found")
    apply_lifecycle_request(db, req)
    req.status = "approved"
    req.review_note = payload.reviewNote
    req.reviewed_by_user_id = user.id
    req.reviewed_at = utcnow()
    db.add(req)
    db.commit()
    return {"request": lifecycle_request_to_dict(req)}


@app.post("/admin/requests/{request_id}/reject")
def reject_request(request_id: str, payload: ReviewRequest, user: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    req = db.scalar(select(StyleLifecycleRequest).where(StyleLifecycleRequest.request_code == request_id))
    if not req:
        raise HTTPException(status_code=404, detail="request not found")
    req.status = "rejected"
    req.review_note = payload.reviewNote
    req.reviewed_by_user_id = user.id
    req.reviewed_at = utcnow()
    db.add(req)
    db.commit()
    return {"request": lifecycle_request_to_dict(req)}


@app.get("/admin/analytics/overview")
def admin_analytics_overview(_: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    counts = aggregate_event_counts(db, days=7)
    style_rows = [style for style in db.scalars(select(Style).order_by(Style.code)).all() if is_real_workspace_style(db, style)]
    style_snapshots = []
    for style in style_rows:
        summary = build_style_trend_summary(db, style.code)
        style_snapshots.append(
            {
                "styleId": style.code,
                "styleName": style.name,
                "healthLabel": summary["scores"]["healthLabel"],
                "compositeRecommendationScore": summary["scores"]["compositeRecommendationScore"],
                "impressions": summary["funnel"]["impressions"],
                "clicks": summary["funnel"]["clicks"],
            }
        )
    return {
        "windowDays": 7,
        "funnel": {
            "impressions": counts["style_impression"],
            "clicks": counts["style_click"],
            "favorites": counts["style_favorite"],
            "tryonStarts": counts["tryon_start"],
            "tryonCompletes": counts["tryon_complete"],
            "bookingCreates": counts["booking_create"],
            "bookingConfirms": counts["booking_confirm"],
        },
        "rates": {
            "clickThroughRate": round(counts["style_click"] / counts["style_impression"], 4) if counts["style_impression"] else 0.0,
            "favoriteRate": round(counts["style_favorite"] / counts["style_click"], 4) if counts["style_click"] else 0.0,
            "tryonStartRate": round(counts["tryon_start"] / counts["style_click"], 4) if counts["style_click"] else 0.0,
            "tryonCompleteRate": round(counts["tryon_complete"] / counts["tryon_start"], 4) if counts["tryon_start"] else 0.0,
            "bookingConversionRate": round(counts["booking_create"] / counts["style_click"], 4) if counts["style_click"] else 0.0,
            "dealConversionRate": round(counts["booking_confirm"] / counts["style_click"], 4) if counts["style_click"] else 0.0,
        },
        "styleSnapshots": style_snapshots,
    }


@app.get("/admin/analytics/events")
def admin_analytics_events(
    _: User = Depends(require_platform_admin),
    db: Session = Depends(get_db),
    event_name: str | None = None,
    style_id: str | None = None,
    limit: int = 100,
) -> dict[str, Any]:
    safe_limit = max(1, min(limit, 500))
    stmt = select(EventLog).order_by(EventLog.occurred_at.desc()).limit(safe_limit)
    if event_name:
        stmt = stmt.where(EventLog.event_name == event_name)
    if style_id:
        stmt = stmt.where(EventLog.style_id == style_id)
    items = db.scalars(stmt).all()
    return {
        "items": [
            {
                "eventId": item.event_id,
                "eventName": item.event_name,
                "userId": item.user_id,
                "deviceId": item.device_id,
                "styleId": item.style_id,
                "storeId": item.store_id,
                "sourcePage": item.source_page,
                "sourceChannel": item.source_channel,
                "sessionId": item.session_id,
                "payload": loads_json(item.payload_json, None),
                "occurredAt": item.occurred_at.isoformat(),
            }
            for item in items
        ]
    }


@app.get("/admin/analytics/styles/{style_id}")
def admin_style_analytics(style_id: str, _: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    style = find_style(db, style_id)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    return {"style": style_to_dict_with_db(db, style), "analytics": build_style_trend_summary(db, style_id)}


@app.get("/admin/trends/recommendations")
def admin_recommendations(_: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    items = db.scalars(select(TrendRecommendation).order_by(TrendRecommendation.created_at.desc())).all()
    return {"items": [recommendation_to_dict(item) for item in items]}


@app.get("/admin/trends/dashboard")
def admin_trends_dashboard(_: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    recommendations = db.scalars(select(TrendRecommendation).order_by(TrendRecommendation.created_at.desc()).limit(12)).all()
    recommendation_items = [recommendation_to_dict(item) for item in recommendations]
    return {
        "community": {
            **build_community_trend_dashboard(db),
            "aiHighlights": build_ai_curated_community_posts(recommendations),
        },
        "product": build_product_trend_dashboard(db),
        "recommendations": recommendation_items,
    }


def serialize_collected_trend_batch(db: Session, topics: list[TrendTopic]) -> dict[str, Any]:
    topic_ids = [topic.id for topic in topics]
    collected_posts = []
    if topic_ids:
        posts = db.scalars(
            select(TrendPost)
            .where(TrendPost.topic_id.in_(topic_ids))
            .order_by(TrendPost.like_count.desc(), TrendPost.collect_count.desc(), TrendPost.comment_count.desc())
        ).all()
        for post in posts:
            meta = load_trend_post_meta(post)
            collected_posts.append(
                {
                    "postId": post.post_id,
                    "url": post.url,
                    "title": post.title,
                    "author": post.author,
                    "imageUrl": meta.get("imageUrl"),
                    "tags": meta.get("tags", []),
                    "likeCount": post.like_count,
                    "collectCount": post.collect_count,
                    "commentCount": post.comment_count,
                    "topicId": post.topic_id,
                }
            )
    return {
        "topicsCollected": len(topics),
        "collectedTopics": [
            {
                "id": topic.id,
                "topicKey": topic.topic_key,
                "title": topic.title,
                "clusterLabel": topic.cluster_label,
                "summary": topic.summary,
                "heatScore": topic.community_heat_score,
                "evidenceCount": topic.evidence_count,
                "lastSeenAt": topic.last_seen_at.isoformat(),
            }
            for topic in topics
        ],
        "collectedPosts": collected_posts,
    }


@app.post("/admin/trends/crawl")
def admin_crawl_trends(
    payload: TrendCollectRequest,
    _: User = Depends(require_platform_admin),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    try:
        collected = collect_xiaohongshu_notes(
            keywords=payload.keywords,
            max_posts_per_keyword=max(1, min(payload.maxPostsPerKeyword, 12)),
            headless=payload.headless,
        )
    except RuntimeError as exc:
        detail = str(exc)
        status_code = 401 if "登录已过期" in detail or "XHS_COOKIES" in detail else 502
        raise HTTPException(status_code=status_code, detail=detail) from exc
    topics = save_xiaohongshu_collection(db, collected)
    batch = serialize_collected_trend_batch(db, topics)
    log_event_json(
        app_logger,
        "trend_crawl_completed",
        keywords=payload.keywords,
        collected=len(collected),
        posts=len(batch["collectedPosts"]),
    )
    return batch


@app.get("/admin/trends/xhs-status")
def admin_xhs_collection_status(
    _: User = Depends(require_platform_admin),
) -> dict[str, Any]:
    return check_xhs_collection_status()


@app.post("/admin/trends/analyze")
def admin_analyze_trends(
    payload: TrendAnalyzeRequest,
    _: User = Depends(require_platform_admin),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    topics = []
    if payload.topicIds:
        topics = db.scalars(select(TrendTopic).where(TrendTopic.id.in_(payload.topicIds))).all()
    created = build_trend_recommendations(db, topics)
    log_event_json(app_logger, "trend_analysis_completed", topicIds=payload.topicIds, recommendationsCreated=created)
    return {
        "topicsAnalyzed": len(topics),
        "recommendationsCreated": created,
        "items": [
            recommendation_to_dict(item)
            for item in db.scalars(select(TrendRecommendation).order_by(TrendRecommendation.created_at.desc()).limit(20)).all()
        ],
    }


@app.post("/admin/trends/collect")
def admin_collect_trends(
    payload: TrendCollectRequest,
    _: User = Depends(require_platform_admin),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    try:
        collected = collect_xiaohongshu_notes(
            keywords=payload.keywords,
            max_posts_per_keyword=max(1, min(payload.maxPostsPerKeyword, 12)),
            headless=payload.headless,
        )
    except RuntimeError as exc:
        detail = str(exc)
        status_code = 401 if "登录已过期" in detail or "XHS_COOKIES" in detail else 502
        raise HTTPException(status_code=status_code, detail=detail) from exc
    topics = save_xiaohongshu_collection(db, collected)
    created = build_trend_recommendations(db, topics)
    topic_ids = [topic.id for topic in topics]
    collected_posts = []
    if topic_ids:
        posts = db.scalars(
            select(TrendPost)
            .where(TrendPost.topic_id.in_(topic_ids))
            .order_by(TrendPost.like_count.desc(), TrendPost.collect_count.desc(), TrendPost.comment_count.desc())
        ).all()
        for post in posts:
            meta = load_trend_post_meta(post)
            collected_posts.append(
                {
                    "postId": post.post_id,
                    "url": post.url,
                    "title": post.title,
                    "author": post.author,
                    "imageUrl": meta.get("imageUrl"),
                    "tags": meta.get("tags", []),
                    "likeCount": post.like_count,
                    "collectCount": post.collect_count,
                    "commentCount": post.comment_count,
                    "topicId": post.topic_id,
                }
            )
    log_event_json(
        app_logger,
        "trend_collect_completed",
        keywords=payload.keywords,
        collected=len(collected),
        recommendationsCreated=created,
    )
    return {
        "topicsCollected": len(topics),
        "recommendationsCreated": created,
        "collectedTopics": [
            {
                "id": topic.id,
                "topicKey": topic.topic_key,
                "title": topic.title,
                "clusterLabel": topic.cluster_label,
                "summary": topic.summary,
                "heatScore": topic.community_heat_score,
                "evidenceCount": topic.evidence_count,
                "lastSeenAt": topic.last_seen_at.isoformat(),
            }
            for topic in topics
        ],
        "collectedPosts": collected_posts,
        "items": [recommendation_to_dict(item) for item in db.scalars(select(TrendRecommendation).order_by(TrendRecommendation.created_at.desc()).limit(20)).all()],
    }


def apply_recommendation(db: Session, rec: TrendRecommendation) -> Style:
    style = find_style(db, rec.target_style_code) if rec.target_style_code else None
    candidate_payload = loads_json(rec.candidate_payload_json, None)
    if rec.recommendation_type == "launch_candidate":
        if not style:
            if not candidate_payload:
                raise HTTPException(status_code=400, detail="recommendation missing candidate payload")
            payload = candidate_payload
            style = Style(
                code=build_style_code(rec.candidate_name or f"trend-{rec.recommendation_code}"),
                name=payload["name"],
                vibe=payload["vibe"],
                price=payload["price"],
                nail_type=payload["nailType"],
                skin_tone=payload["skinTone"],
                tags_json=dumps_json(payload.get("tags", [])),
                colors_json=dumps_json(payload.get("colors", [])),
                status="active",
                updated_at=utcnow(),
            )
            db.add(style)
            db.flush()
            rec.target_style_code = style.code
            if isinstance(candidate_payload, dict) and candidate_payload.get("imageUrl"):
                asset = ensure_style_asset_record(db, style)
                asset.enhanced_url = candidate_payload["imageUrl"]
                asset.display_name = style.name
                asset.tags_json = style.tags_json
                asset.updated_at = utcnow()
                db.add(asset)
        else:
            style.status = "active"
            style.updated_at = utcnow()
    elif rec.recommendation_type == "boost_candidate":
        if not style:
            raise HTTPException(status_code=400, detail="recommendation has no target style")
        style.status = "active"
        style.updated_at = utcnow()
    elif rec.recommendation_type in {"deprioritize_candidate", "delist_candidate"}:
        if not style:
            raise HTTPException(status_code=400, detail="recommendation has no target style")
        style.status = "inactive" if rec.recommendation_type == "delist_candidate" else "draft"
        style.updated_at = utcnow()
    else:
        raise HTTPException(status_code=400, detail="unsupported recommendation type")
    if rec.target_store_code and style:
        listing = get_or_create_listing(db, rec.target_store_code, style.code, style.price)
        listing.status = "active" if rec.recommendation_type in {"launch_candidate", "boost_candidate"} else "inactive"
        listing.published_at = listing.published_at or utcnow()
    return style


@app.post("/admin/trends/recommendations/{recommendation_id}/approve")
def approve_recommendation(recommendation_id: str, payload: ReviewRequest, user: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    rec = db.scalar(select(TrendRecommendation).where(TrendRecommendation.recommendation_code == recommendation_id))
    if not rec:
        raise HTTPException(status_code=404, detail="recommendation not found")
    style = apply_recommendation(db, rec)
    rec.status = "approved"
    rec.reviewed_by_user_id = user.id
    rec.reviewed_at = utcnow()
    rec.action_text = f"{rec.action_text} 审核备注：{payload.reviewNote}".strip()
    db.add(rec)
    db.commit()
    return {"recommendation": recommendation_to_dict(rec), "style": style_to_dict_with_db(db, style)}


@app.post("/admin/trends/recommendations/{recommendation_id}/reject")
def reject_recommendation(recommendation_id: str, payload: ReviewRequest, user: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    rec = db.scalar(select(TrendRecommendation).where(TrendRecommendation.recommendation_code == recommendation_id))
    if not rec:
        raise HTTPException(status_code=404, detail="recommendation not found")
    rec.status = "rejected"
    rec.reviewed_by_user_id = user.id
    rec.reviewed_at = utcnow()
    rec.action_text = f"{rec.action_text} 审核备注：{payload.reviewNote}".strip()
    db.add(rec)
    db.commit()
    return {"recommendation": recommendation_to_dict(rec)}


@app.post("/admin/trends/import")
def admin_import_trend(payload: TrendImportRequest, _: User = Depends(require_platform_admin), db: Session = Depends(get_db)) -> dict[str, Any]:
    topic_key = build_style_code(f"{payload.clusterLabel}-{payload.title}")[:40]
    topic = TrendTopic(
        topic_key=topic_key,
        platform="xiaohongshu",
        title=payload.title,
        cluster_label=payload.clusterLabel,
        summary=payload.summary,
        community_heat_score=payload.communityHeatScore,
        evidence_count=1,
        last_seen_at=utcnow(),
    )
    db.add(topic)
    db.flush()
    db.add(
        TrendRecommendation(
            recommendation_code=next_recommendation_code(db),
            recommendation_type=payload.recommendationType,
            target_style_code=payload.targetStyleId,
            candidate_name=payload.candidateName or payload.title,
            trigger_reason=payload.summary,
            community_evidence="人工导入的小红书主题，待运营确认。",
            in_app_evidence="站内数据将在下一轮聚合中补齐。",
            confidence_score=round(min(payload.communityHeatScore / 100, 0.99), 2),
            action_text="人工导入后待审核。",
            prerequisites="需要运营确认门店供给。",
        )
    )
    db.commit()
    return {"topicKey": topic.topic_key}


@app.get("/admin/merchants/me/dashboard")
def merchant_dashboard(user: User = Depends(require_merchant_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    if not user.managed_store_code:
        return {
            "store": None,
            "summary": {"bookingCount": 0, "pendingRequestCount": 0, "listingCount": 0},
        }
    store = find_store(db, user.managed_store_code)
    if not store:
        return {
            "store": None,
            "summary": {"bookingCount": 0, "pendingRequestCount": 0, "listingCount": 0},
        }
    booking_count = db.query(Booking).filter(Booking.store_id == store.code).count()
    pending_requests = db.query(StyleLifecycleRequest).filter(StyleLifecycleRequest.store_code == store.code, StyleLifecycleRequest.status == "pending").count()
    all_listings = db.scalars(select(StoreStyleListing).where(StoreStyleListing.store_code == store.code)).all()
    listings = sum(1 for item in all_listings if is_real_workspace_style(db, find_style(db, item.style_code)))
    return {
        "store": store_to_dict(store),
        "summary": {
            "bookingCount": booking_count,
            "pendingRequestCount": pending_requests,
            "listingCount": listings,
        },
    }


@app.get("/admin/merchants/me/listings")
def merchant_listings(user: User = Depends(require_merchant_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    if not user.managed_store_code:
        return {"items": []}
    listings = db.scalars(select(StoreStyleListing).where(StoreStyleListing.store_code == user.managed_store_code).order_by(StoreStyleListing.style_code)).all()
    payload = []
    for item in listings:
        style = find_style(db, item.style_code)
        if not is_real_workspace_style(db, style):
            continue
        payload.append(listing_to_dict_with_style(db, item))
    return {"items": payload}


@app.patch("/admin/merchants/me/listings/{listing_id}")
def merchant_update_listing(listing_id: int, payload: MerchantListingUpdateRequest, user: User = Depends(require_merchant_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    listing = db.get(StoreStyleListing, listing_id)
    if not listing:
        raise HTTPException(status_code=404, detail="listing not found")
    require_store_access(user, listing.store_code)
    if payload.price is not None:
        listing.price = payload.price
    if payload.inventoryCount is not None:
        listing.inventory_count = payload.inventoryCount
    if payload.status is not None:
        listing.status = payload.status
    listing.updated_at = utcnow()
    db.add(listing)
    db.commit()
    return {"listing": listing_to_dict(listing)}


@app.patch("/admin/merchants/me/store")
def merchant_update_store(payload: MerchantStoreUpdateRequest, user: User = Depends(require_merchant_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    if not user.managed_store_code:
        raise HTTPException(status_code=400, detail="merchant user is not bound to a store")
    store = find_store(db, user.managed_store_code)
    if not store:
        raise HTTPException(status_code=404, detail="store not found")
    if payload.slots is not None:
        store.slots_json = dumps_json(payload.slots)
    if payload.isAcceptingBookings is not None:
        store.is_accepting_bookings = payload.isAcceptingBookings
    store.updated_at = utcnow()
    db.add(store)
    db.commit()
    return {"store": store_to_dict(store)}


@app.get("/admin/merchants/me/bookings")
def merchant_bookings(user: User = Depends(require_merchant_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    if not user.managed_store_code:
        return {"items": []}
    bookings = db.scalars(select(Booking).where(Booking.store_id == user.managed_store_code).order_by(Booking.created_at.desc())).all()
    return {"items": [booking_response(item) for item in bookings]}


@app.get("/admin/merchants/me/requests")
def merchant_requests(user: User = Depends(require_merchant_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    if not user.managed_store_code:
        return {"items": []}
    items = db.scalars(
        select(StyleLifecycleRequest)
        .where(StyleLifecycleRequest.store_code == user.managed_store_code)
        .order_by(StyleLifecycleRequest.created_at.desc())
    ).all()
    payload = []
    for item in items:
        style = find_style(db, item.style_code) if item.style_code else None
        if not is_real_workspace_style(db, style):
            continue
        request_payload = lifecycle_request_to_dict(item)
        request_payload["styleName"] = style.name if style else item.style_code
        payload.append(request_payload)
    return {"items": payload}


@app.post("/admin/merchants/me/requests")
def merchant_create_request(payload: MerchantLifecycleRequestCreate, user: User = Depends(require_merchant_user), db: Session = Depends(get_db)) -> dict[str, Any]:
    store_code = payload.storeId or user.managed_store_code
    if not store_code:
        raise HTTPException(status_code=400, detail="storeId is required")
    require_store_access(user, store_code)
    style = find_style(db, payload.styleId)
    if not style:
        raise HTTPException(status_code=404, detail="style not found")
    req = StyleLifecycleRequest(
        request_code=next_request_code(db),
        requested_by_user_id=user.id,
        merchant_id=user.merchant_id,
        store_code=store_code,
        style_code=payload.styleId,
        requested_action=payload.requestedAction,
        reason=payload.reason,
        status="pending",
    )
    db.add(req)
    db.commit()
    db.refresh(req)
    return {"request": lifecycle_request_to_dict(req)}


@app.post("/internal/try-on/jobs/claim")
def worker_claim_job(_: None = Depends(require_worker), db: Session = Depends(get_db)) -> dict[str, Any]:
    job = db.scalar(select(TryOnJob).where(TryOnJob.status == "queued").order_by(TryOnJob.created_at.asc()))
    if not job:
        return {"job": None}
    style = find_style(db, job.style_id)
    if not style:
        job.status = "failed"
        job.error_code = "STYLE_NOT_FOUND"
        job.error_message = "style not found"
        db.add(job)
        db.commit()
        return {"job": None}
    job.status = "processing"
    job.stage = "preparing"
    job.progress = 15
    job.claimed_by_worker = True
    job.updated_at = utcnow()
    result_image_key = f"{job.job_code}.png"
    db.add(job)
    db.commit()
    return {
        "job": {
            "id": job.job_code,
            "styleId": style.code,
            "styleName": style.name,
            "sourceImagePath": str(settings.uploads_dir / job.source_image_key),
            "resultImagePath": str(settings.results_dir / result_image_key),
            "resultImageKey": result_image_key,
            "selectedLength": job.selected_length,
            "selectedShape": job.selected_shape,
            "styleColors": loads_json(style.colors_json, []),
        }
    }


@app.post("/internal/try-on/jobs/{job_code}/progress")
def worker_progress(job_code: str, payload: WorkerProgressRequest, _: None = Depends(require_worker), db: Session = Depends(get_db)) -> dict[str, Any]:
    job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code))
    if not job:
        raise HTTPException(status_code=404, detail="try-on job not found")
    job.status = "processing"
    job.stage = payload.stage
    job.progress = payload.progress
    job.updated_at = utcnow()
    db.add(job)
    db.commit()
    return {"status": "ok"}


@app.post("/internal/try-on/jobs/{job_code}/complete")
def worker_complete(job_code: str, payload: WorkerCompleteRequest, _: None = Depends(require_worker), db: Session = Depends(get_db)) -> dict[str, Any]:
    job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code))
    if not job:
        raise HTTPException(status_code=404, detail="try-on job not found")
    job.status = "completed"
    job.stage = "completed"
    job.progress = 100
    job.result_image_key = payload.resultImageKey
    job.detected_traits = serialize_traits(payload.detectedTraits)
    job.completed_at = utcnow()
    job.updated_at = utcnow()
    db.add(job)
    asset = get_style_asset_by_code(db, job.style_id)
    if asset:
        db.add(
            TryOnRecord(
                user_id=job.user_id,
                hand_image_id=None,
                nail_style_asset_id=asset.id,
                result_url=tryon_result_url(payload.resultImageKey),
                source="async-job",
                duration_ms=0,
                selected_length=job.selected_length,
                selected_shape=job.selected_shape,
            )
        )
    log_event(db, "tryon_complete", user_id=job.user_id, style_id=job.style_id, source_page="tryon_worker")
    db.commit()
    log_event_json(app_logger, "tryon_async_completed", jobId=job.job_code, userId=job.user_id, styleId=job.style_id, resultImageKey=job.result_image_key)
    return {"status": "ok"}


@app.post("/internal/try-on/jobs/{job_code}/fail")
def worker_fail(job_code: str, payload: WorkerFailRequest, _: None = Depends(require_worker), db: Session = Depends(get_db)) -> dict[str, Any]:
    job = db.scalar(select(TryOnJob).where(TryOnJob.job_code == job_code))
    if not job:
        raise HTTPException(status_code=404, detail="try-on job not found")
    job.status = "failed"
    job.stage = "failed"
    job.error_code = payload.errorCode
    job.error_message = payload.errorMessage
    job.updated_at = utcnow()
    db.add(job)
    db.commit()
    return {"status": "ok"}


@app.post("/internal/trends/run")
def run_trend_refresh(_: None = Depends(require_worker), db: Session = Depends(get_db)) -> dict[str, Any]:
    topics = db.scalars(select(TrendTopic).order_by(TrendTopic.community_heat_score.desc())).all()
    created = build_trend_recommendations(db, list(topics))
    log_event_json(app_logger, "trend_refresh_completed", created=created)
    return {"created": created}
