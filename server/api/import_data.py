from __future__ import annotations

import shutil
import sys
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook
from sqlalchemy import select

sys.path.insert(0, str(Path(__file__).resolve().parent))

from app.database import SessionLocal, Base, engine  # noqa: E402
from app.main import (  # noqa: E402
    settings,
    utcnow,
)
from app.models import HandImage, NailStyleAsset, Style  # noqa: E402
from app.domain import dumps_json  # noqa: E402


def file_ext_from_url(url: str, default: str = ".png") -> str:
    suffix = Path(urlparse(url).path).suffix.lower()
    return suffix if suffix in {".png", ".jpg", ".jpeg", ".webp"} else default


def download(url: str, target: Path) -> bool:
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(response.content)
        return True
    except Exception:
        return False


def main() -> None:
    workbook_name = "命题三美甲评测数据（对外版）.xlsx"
    candidates = [
        Path(__file__).resolve().parents[1] / workbook_name,
        Path(__file__).resolve().parent / workbook_name,
        Path.cwd() / workbook_name,
        Path("/") / workbook_name,
    ]
    workbook = next((path for path in candidates if path.exists()), None)
    if workbook is None:
        workbook = candidates[0]
    if not workbook.exists():
        raise SystemExit(f"workbook not found: {workbook}")

    settings.static_styles_dir.mkdir(parents=True, exist_ok=True)
    settings.static_hands_dir.mkdir(parents=True, exist_ok=True)
    Base.metadata.create_all(bind=engine)

    wb = load_workbook(workbook, read_only=True, data_only=True)
    style_rows = list(wb["款式图"].iter_rows(values_only=True))[1:]
    hand_rows = list(wb["手图"].iter_rows(values_only=True))[1:]

    with SessionLocal() as db:
        for idx, row in enumerate(style_rows, start=1):
            style_code = f"style-{idx:02d}"
            original_url = str(row[1]) if len(row) > 1 and row[1] else None
            enhanced_url = str(row[2]) if len(row) > 2 and row[2] else None
            local_path = settings.static_styles_dir / f"style_{idx:02d}{file_ext_from_url(enhanced_url or original_url or '')}"
            if enhanced_url:
                download(enhanced_url, local_path)

            asset = db.scalar(select(NailStyleAsset).where(NailStyleAsset.style_code == style_code))
            if not asset:
                asset = NailStyleAsset(style_code=style_code, display_name=f"款式 {idx:02d}", sequence_no=idx)
                db.add(asset)
            asset.original_url = original_url
            asset.enhanced_url = enhanced_url
            asset.local_image_path = str(local_path) if local_path.exists() else None
            asset.category = ""
            asset.color_tone = ""
            asset.tags_json = dumps_json([])
            asset.price = ""
            asset.popularity = 0
            asset.updated_at = utcnow()

            style = db.scalar(select(Style).where(Style.code == style_code))
            if not style:
                style = Style(code=style_code, name=asset.display_name, vibe="", price="", nail_type="", skin_tone="")
                db.add(style)
            style.name = asset.display_name
            style.vibe = ""
            style.price = ""
            style.nail_type = ""
            style.skin_tone = ""
            style.tags_json = dumps_json([])
            style.colors_json = dumps_json([])
            style.status = "active"
            style.updated_at = utcnow()

        seen_hands: set[str] = set()
        hand_index = 0
        for row in hand_rows:
            if not row or not row[0]:
                continue
            hand_url = str(row[0])
            if hand_url in seen_hands:
                continue
            seen_hands.add(hand_url)
            hand_index += 1
            hand_code = f"hand_{hand_index:02d}"
            local_path = settings.static_hands_dir / f"{hand_code}{file_ext_from_url(hand_url)}"
            download(hand_url, local_path)

            hand = db.scalar(select(HandImage).where(HandImage.hand_code == hand_code))
            if not hand:
                hand = HandImage(hand_code=hand_code, image_url=hand_url)
                db.add(hand)
            hand.local_path = str(local_path) if local_path.exists() else None
            hand.source_type = "preset"
            hand.skin_tone = ""
            hand.hand_type = ""

        db.commit()

    print(f"Imported {len(style_rows)} styles and {hand_index} unique hands from {workbook.name}")


if __name__ == "__main__":
    main()
